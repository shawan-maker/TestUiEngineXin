#!/usr/bin/env python3
"""validate_excel.py — Excel 测试用例预检工具 (Phase 1)

L1 规则（23条）: 全自动修复 [v3: +R23-R29, R31-R34, R36, RQ]
L2 规则（7条）: 自动修复 + 标记待确认

退出码:
  0 = 无问题或仅有 info 级提示
  1 = 有 L1/L2 自动修复
  2 = 运行失败

用法:
  python validate_excel.py input.xlsx
  python validate_excel.py input.xlsx --output corrected.xlsx --report report.html
"""

import argparse
import json
import os
import re
import sys
from copy import copy
from pathlib import Path
from typing import Dict, List, Tuple

# Ensure tools/ is on sys.path for core.* imports
_tools_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _tools_dir not in sys.path:
    sys.path.insert(0, _tools_dir)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(2)

# Ensure tools/ is on sys.path for core.* imports
_tools_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _tools_dir not in sys.path:
    sys.path.insert(0, _tools_dir)


# ─── 数据结构 ───

class FixRecord:
    """一条修复记录"""
    def __init__(self, level: str, rule_id: str, sheet: str,
                 case_name: str, step_num: int, before: str, after: str,
                 auto_fixed: bool = True,
                 severity: str = None, suggestion: str = None):
        self.level = level        # L1 / L2 / L3
        self.rule_id = rule_id    # R01-R36, RQ
        self.sheet = sheet
        self.case_name = case_name
        self.step_num = step_num
        self.before = before
        self.after = after
        self.auto_fixed = auto_fixed  # True=已自动修复, False=待确认
        self.severity = severity      # 'error' | 'warn' | None
        self.suggestion = suggestion  # 修改建议（L3 error 时提供）

    def to_dict(self):
        return {
            'level': self.level, 'rule_id': self.rule_id,
            'sheet': self.sheet, 'case_name': self.case_name,
            'step_num': self.step_num, 'before': self.before,
            'after': self.after, 'auto_fixed': self.auto_fixed,
            'severity': self.severity, 'suggestion': self.suggestion,
        }


class ExcelValidator:
    def __init__(self, input_path: str, output_path: str = None,
                 report_path: str = None):
        self.input_path = Path(input_path)
        if not self.input_path.exists():
            print(f"ERROR: File not found: {input_path}")
            sys.exit(2)

        stem = self.input_path.stem
        parent = self.input_path.parent

        self.output_path = output_path or str(parent / f"{stem}-修正版.xlsx")
        self.report_path = report_path or str(parent / "excel_validation_report.html")

        self.wb = openpyxl.load_workbook(input_path)
        self.fixes: List[FixRecord] = []
        self.stats = {'L1_auto': 0, 'L2_auto': 0, 'L2_warn': 0, 'L3_error': 0}

    # ─── 主入口 ───

    def run(self) -> int:
        """执行所有检查，返回 exit code

        退出码:
          0 = 通过（无需清洗 或 L1/L2 已自动修复）
          1 = L3 解析验证失败（阻断，需修改 Excel 后重跑）
        """
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            self._process_sheet(ws, sheet_name)

        # 分类结果
        l3_errors = [f for f in self.fixes
                     if f.level == 'L3' and f.severity == 'error']
        l1_l2_fixes = [f for f in self.fixes
                       if f.level in ('L1', 'L2') and f.auto_fixed]

        # 生成报告（始终生成，方便审查）
        self._generate_report()

        # 三路出口
        if l3_errors:
            # L3 阻断: 保存 L1/L2 已修复部分 + 问题清单
            self.wb.save(self.output_path)
            print(f"\n{'=' * 60}")
            print(f"L3 PARSE VALIDATION FAILED - {len(l3_errors)} unparseable steps")
            print(f"  L1 auto-fixed: {self.stats['L1_auto']}")
            print(f"  L2 auto-fixed: {self.stats['L2_auto']}")
            print(f"  L3 errors:     {len(l3_errors)}")
            print(f"{'=' * 60}")
            print(f"Corrected Excel (L1/L2 fixes applied): {self.output_path}")
            print(f"Report: {self.report_path}")
            print(f"Please fix the L3 errors in Excel and re-run.")
            return 1
        elif l1_l2_fixes:
            # L3 通过，有 L1/L2 自动修复 → 生成修正版
            self.wb.save(self.output_path)
            print(f"\n{'=' * 60}")
            print(f"Excel precheck done - {len(l1_l2_fixes)} issues auto-fixed")
            print(f"  L1 auto-fixed: {self.stats['L1_auto']}")
            print(f"  L2 auto-fixed: {self.stats['L2_auto']}")
            print(f"  L2 need-review: {self.stats.get('L2_warn', 0)}")
            print(f"{'=' * 60}")
            print(f"Corrected: {self.output_path}")
            print(f"Report:    {self.report_path}")
            print(f"后续流程请使用修正版文件。")
            return 0
        else:
            # L3 通过，L1/L2 零修复 → 不生成副本，直接用原始 Excel
            print(f"\n{'=' * 60}")
            print(f"Excel precheck passed - no issues found")
            print(f"{'=' * 60}")
            print(f"No corrections needed. Use original file: {self.input_path}")
            print(f"Report: {self.report_path}")
            print(f"后续流程直接使用原始文件。")
            return 0

    # ─── Sheet 处理 ───

    def _process_sheet(self, ws, sheet_name: str):
        """处理单个 sheet 的所有用例"""
        headers = [cell.value for cell in ws[1]]
        steps_col = self._find_steps_column(headers)
        name_col = self._find_name_column(headers)

        if steps_col is None:
            return

        for row_idx in range(2, ws.max_row + 1):
            steps_val = ws.cell(row=row_idx, column=steps_col).value
            name_val = ws.cell(row=row_idx, column=name_col).value if name_col else ""

            if not steps_val or not isinstance(steps_val, str):
                continue

            case_name = name_val or f"Row-{row_idx}"
            original = steps_val

            # 依次应用所有规则
            steps_val = self._apply_l1_rules(steps_val, sheet_name, case_name, row_idx)
            steps_val = self._apply_l2_rules(steps_val, sheet_name, case_name, row_idx)

            # R20: 复合步骤拆分（L1+L2 产生的"操作+等待"合并步骤）
            steps_val = self._r20_split_compound_steps(steps_val, sheet_name, case_name)

            # R20b: 去重连续相同的等待步骤
            steps_val = self._dedup_consecutive_waits(steps_val, sheet_name, case_name)

            # L3a: 验证清洗后的步骤是否可被 StepParser 解析
            self._apply_l3_parse_validation(steps_val, sheet_name, case_name)

            # L3b/c: L3 关键字白名单 + 参数校验
            self._apply_l3_whitelist_validation(steps_val, sheet_name, case_name)

            # 如果有修改，回写
            if steps_val != original:
                ws.cell(row=row_idx, column=steps_col).value = steps_val

    def _find_steps_column(self, headers: list) -> int:
        """找到步骤列（1-indexed）"""
        for i, h in enumerate(headers, 1):
            if h and ('步骤' in str(h) or '内容' in str(h)):
                return i
        return 3  # 默认第3列

    def _find_name_column(self, headers: list) -> int:
        """找到用例名称列（1-indexed）"""
        for i, h in enumerate(headers, 1):
            if h and ('名称' in str(h) or '用例名' in str(h)):
                return i
        return 2  # 默认第2列

    # ─── 步骤解析 ───

    def _parse_steps(self, text: str) -> List[Tuple[int, str]]:
        """解析步骤文本为 [(step_num, content), ...]
        只匹配行首的 "数字. " 模式，避免 URL 中的数字被误识别
        """
        steps = []
        current_num = None
        current_parts = []

        for line in text.split('\n'):
            line_stripped = line.strip()
            if not line_stripped:
                continue

            # 只在行首匹配步骤编号
            m = re.match(r'^(\d+)\s*[.、．]\s*(.*)', line_stripped)
            if m:
                # 保存前一个步骤
                if current_num is not None:
                    steps.append((current_num, '\n'.join(current_parts).strip()))
                current_num = int(m.group(1))
                current_parts = [m.group(2)]
            elif current_num is not None:
                # 续行（属于当前步骤的内容）
                current_parts.append(line_stripped)

        # 保存最后一个步骤
        if current_num is not None:
            steps.append((current_num, '\n'.join(current_parts).strip()))

        return steps

    def _rebuild_steps(self, steps: List[Tuple[int, str]]) -> str:
        """将步骤列表重建为文本"""
        lines = []
        for i, (_, content) in enumerate(steps, 1):
            lines.append(f"{i}. {content}")
        return '\n'.join(lines)

    def _record_fix(self, level: str, rule_id: str, sheet: str,
                    case_name: str, step_num: int, before: str, after: str,
                    auto_fixed: bool = True,
                    severity: str = None, suggestion: str = None):
        """记录修复"""
        rec = FixRecord(level, rule_id, sheet, case_name, step_num,
                        before, after, auto_fixed, severity, suggestion)
        self.fixes.append(rec)
        if severity == 'error':
            self.stats['L3_error'] = self.stats.get('L3_error', 0) + 1
        elif auto_fixed:
            self.stats[f'{level}_auto'] = self.stats.get(f'{level}_auto', 0) + 1
        else:
            self.stats[f'{level}_warn'] = self.stats.get(f'{level}_warn', 0) + 1

    # ─── L1 规则（全自动修复） ───

    def _apply_l1_rules(self, text: str, sheet: str, case_name: str,
                        row_idx: int) -> str:
        """应用 L1 规则，返回修改后的文本 [v3 执行顺序]

        v3 新增: R23-R29, R31-R34, R36 + _enforce_quotes
        v3 执行顺序: R31(去注释) → R33(无前缀下拉) → R36(找到→点击) → R32(返回)
                     → R23(断言) → R24(自定义关键词) → R26(输入框) → R27_v3(等待)
                     → R28(导航) → R29(下拉框修复)
        """
        steps = self._parse_steps(text)
        if not steps:
            return text

        modified = False
        new_steps = []

        for step_num, content in steps:
            original_content = content

            # ─── 现有规则（v3-audit: R05 提前到 R02 之前）───
            # R01: 断言关键词统一
            content = self._r01_assertion_keyword(content, sheet, case_name, step_num)
            # R05: 引号统一（所有引号变体 → ASCII ""）[B1] [v3-audit: 必须在 R02 之前]
            content = self._r05_quote_normalize(content, sheet, case_name, step_num)
            # R02: 数据清洗（\t \n 多余空格）[依赖 R05 先统一引号]
            content = self._r02_data_clean(content, sheet, case_name, step_num)
            # R07: 查询步骤标准化
            content = self._r07_query_standardize(content, sheet, case_name, step_num)
            # R08: 结束句标准化
            content = self._r08_end_sentence(content, sheet, case_name, step_num)
            # R09: URL 空格
            content = self._r09_url_space(content, sheet, case_name, step_num)
            # R10: 按钮空格去除 [B3]
            content = self._r10_button_normalize(content, sheet, case_name, step_num)
            # R18: 步骤格式标准化（排除集含弯引号 [B2]）
            content = self._r18_step_normalize(content, sheet, case_name, step_num)
            # R21: 过滤注释续行（# 或 // 开头）
            content = self._r21_filter_comments(content, sheet, case_name, step_num)
            # R22: 过滤分隔符行（--- / === / ***）
            content = self._r22_filter_separators(content, sheet, case_name, step_num)

            # ─── 新增规则 (v3) ───
            # R31: 去注释（最先运行，避免注释干扰后续规则）
            content = self._r31_inline_comment_strip(content, sheet, case_name, step_num)
            # R33: 无前缀下拉框修复
            content = self._r33_dropdown_no_prefix_fix(content, sheet, case_name, step_num)
            # R36: 找到XX区域 → 点击XX区域
            content = self._r36_find_to_click(content, sheet, case_name, step_num)
            # R32: 返回页面 → 返回
            content = self._r32_go_back_normalize(content, sheet, case_name, step_num)
            # R23: 断言格式统一
            content = self._r23_assertion_format(content, sheet, case_name, step_num)
            # R24: 自定义关键词转换
            content = self._r24_custom_keyword_convert(content, sheet, case_name, step_num)
            # R26: 输入框格式补齐
            content = self._r26_input_field_normalize(content, sheet, case_name, step_num)
            # R27: 等待格式标准化（仅模式1+4）
            content = self._r27_wait_normalize(content, sheet, case_name, step_num)
            # R28: 导航/跳转格式
            content = self._r28_navigate_normalize(content, sheet, case_name, step_num)
            # R29: 下拉框操作修复
            content = self._r29_dropdown_fix(content, sheet, case_name, step_num)

            if content != original_content:
                modified = True

            new_steps.append((step_num, content))

        # R03: 步骤编号连续性（在所有步骤处理完后检查）
        rebuilt = self._rebuild_steps(new_steps)
        if rebuilt != text:
            nums = [s[0] for s in new_steps]
            expected = list(range(1, len(nums) + 1))
            if nums != expected:
                self._record_fix('L1', 'R03', sheet, case_name, 0,
                                 f"步骤编号: {nums}", f"重编号为: {expected}")
                rebuilt = self._rebuild_steps([(i, s[1]) for i, s in
                                               enumerate(new_steps, 1)])

        # ─── 步骤级规则（可能拆分/新增步骤）───
        # R04: 合并步骤拆分
        rebuilt = self._r04_split_merged(rebuilt, sheet, case_name)
        # R06: 断言拆分
        rebuilt = self._r06_split_assertions(rebuilt, sheet, case_name)
        # R25: 等待+断言/条件 复合拆分
        rebuilt = self._r25_split_wait_assertion(rebuilt, sheet, case_name)
        # R34: 操作+等待+条件 三段复合拆分
        rebuilt = self._r34_split_action_wait_condition(rebuilt, sheet, case_name)

        # ─── 引号强制执行（L1 最后一步）───
        # R37: 引号完整性 — 已分类步骤的操作目标补引号（补充 _enforce_quotes 的缺口）
        rebuilt = self._r37_quote_integrity(rebuilt, sheet, case_name)
        rebuilt = self._enforce_quotes(rebuilt, sheet, case_name)

        return rebuilt

    def _r01_assertion_keyword(self, content: str, sheet: str,
                               case_name: str, step_num: int) -> str:
        """R01: 断言关键词统一 — 文本断言 → 断言"""
        # 只处理明确的 "文本断言" 模式
        m = re.match(r'^文本断言[：:]\s*(.*)', content)
        if m:
            rest = m.group(1)
            new_content = f"断言：页面提示{rest}" if '页面提示' not in rest else f"断言：{rest}"
            self._record_fix('L1', 'R01', sheet, case_name, step_num,
                             content, new_content)
            return new_content
        return content

    def _r02_data_clean(self, content: str, sheet: str,
                        case_name: str, step_num: int) -> str:
        """R02: 数据清洗 — 去除 \\t \\n 和多余空格

        改进：控制字符替换为空格（而非删除），避免丢失引号内的文本内容。
        例: "工单标题"\\t\\n肖洋发起的制品出库" → "工单标题" 肖洋发起的制品出库"
        """
        new = content
        # 去除字面量 \t 和 \n（Excel 单元格中实际输入的转义字符）
        # 匹配引号后的 \t \n（引号包括 ASCII “ ' 和弯引号）
        new = re.sub(r'(?<=[\x22\x27])\\t', '', new)
        new = re.sub(r'(?<=[\x22\x27])\\n', '', new)
        # 替换实际的控制字符为空格（保留引号内文本内容）
        new = re.sub(r'(?<=[\x22\x27])\t+', ' ', new)
        new = re.sub(r'(?<=[\x22\x27])\n+', ' ', new)
        # \t\n 组合替换为单个空格（不删除，避免丢失文本）
        new = new.replace('\t\n', ' ')
        new = new.replace('\t', ' ')
        # 合并多余空格
        new = re.sub(r'  +', ' ', new)
        # 去除值前后的多余空格（引号内侧）
        new = re.sub(r'([\x22\x27])\s+(?=[^\s\x22\x27])', r'\1', new)
        new = re.sub(r'(?<=[^\s\x22\x27])\s+([\x22\x27])', r'\1', new)

        if new != content:
            self._record_fix('L1', 'R02', sheet, case_name, step_num,
                             repr(content), repr(new))
        return new

    def _r05_quote_normalize(self, content: str, sheet: str,
                             case_name: str, step_num: int) -> str:
        """R05: 引号统一 — '' → ""，'' → ""，「」 → "" """
        new = content
        # ASCII 单引号 'XX' → "XX"（仅在步骤指令上下文中）
        new = re.sub(r"(?<!\w)'([^']+)'(?!\w)", r'"\1"', new)
        # [v3] 弯单引号 '' → ASCII 双引号 ""
        new = re.sub(r"\u2018([^\u2019]+)\u2019", r'"\1"', new)
        # [v3] 角引号 「」 → 弯双引号 ""（Q 字符类已覆盖弯双引号）
        new = new.replace('「', '\u201c').replace('」', '\u201d')
        # [v3-audit] Normalize curly double quotes to ASCII double quotes
        new = new.replace('“', '"').replace('”', '"')

        if new != content:
            self._record_fix('L1', 'R05', sheet, case_name, step_num,
                             content, new)
        return new

    def _r07_query_standardize(self, content: str, sheet: str,
                               case_name: str, step_num: int) -> str:
        """R07: 查询步骤标准化"""
        patterns = [
            (r"^点击['\"]?查询['\"]?按钮进行搜索$",
             '点击"查询"按钮，等待查询结果展示成功'),
            (r"^点击['\"]?查询['\"]?按钮，等待搜索结果$",
             '点击"查询"按钮，等待查询结果展示成功'),
            (r"^点击['\"]?查询['\"]?按钮，等待搜索结果加载完成$",
             '点击"查询"按钮，等待查询结果展示成功'),
            (r"^点击['\"]?查询['\"]?按钮$",
             '点击"查询"按钮，等待查询结果展示成功'),
        ]
        for pat, replacement in patterns:
            if re.match(pat, content):
                self._record_fix('L1', 'R07', sheet, case_name, step_num,
                                 content, replacement)
                return replacement
        return content

    def _r08_end_sentence(self, content: str, sheet: str,
                          case_name: str, step_num: int) -> str:
        """R08: 结束句标准化"""
        standard = "请完成测试结论的总结并结束任务"
        variants = [
            r"完成测试结论的总结并结束任务",
            r"请完成测试结论总结并结束",
            r"结束任务",
        ]
        for v in variants:
            if re.match(f'^{v}$', content) and content != standard:
                self._record_fix('L1', 'R08', sheet, case_name, step_num,
                                 content, standard)
                return standard
        return content

    def _r09_url_space(self, content: str, sheet: str,
                       case_name: str, step_num: int) -> str:
        """R09: URL 空格 — 访问http:// → 访问 http://"""
        m = re.match(r'^访问(http://.+)$', content)
        if m:
            new = f"访问 {m.group(1)}"
            self._record_fix('L1', 'R09', sheet, case_name, step_num,
                             content, new)
            return new
        return content

    def _r10_button_normalize(self, content: str, sheet: str,
                              case_name: str, step_num: int) -> str:
        """R10: 按钮空格去除 — "确 定" → "确定" [v3 实现]"""
        buttons = ['确 定', '取 消', '保 存', '提 交', '确 认', '搜 索',
                   '查 询', '新 增', '删 除', '编 辑', '导 出']
        new = content
        for btn in buttons:
            new = new.replace(btn, btn.replace(' ', ''))
        if new != content:
            self._record_fix('L1', 'R10', sheet, case_name, step_num,
                             content, new)
        return new

    def _r18_step_normalize(self, content: str, sheet: str,
                            case_name: str, step_num: int) -> str:
        """R18: 步骤格式标准化 — 将常见变体统一为 StepParser 可解析格式

        保证清洗后的步骤能被 step_patterns.py 的模式匹配，
        减少 _case_generator.py 中的 AI 兜底。
        """
        new = content

        # 18a: 无引号的字段名加引号
        #   在项目名称下拉框选择XX → 在"项目名称"下拉框中选择"XX"
        # [v3] 排除集扩展: 包含所有引号变体，已有引号的字段名不被再加一层
        m = re.match(r"^在([^\s\"""“”'‘’]{2,10}?)((?:下拉框|输入框|文本框|时间选择框|(?<!级联)(?<!时间)选择框|框)"
                      r"(?:中)?(?:输入|选择).+)$", new)
        if m:
            field, rest = m.group(1), m.group(2)
            # 如果字段名不包含特殊字符且看起来像中文标签
            if re.match(r'^[一-鿿\w]+$', field):
                new = f'在"{field}"{rest}'
                self._record_fix('L1', 'R18', sheet, case_name, step_num,
                                 content, new)

        # 18b: 统一"下拉框选择"为"下拉框中选择"
        new2 = re.sub(
            r'(在["\u201c][^"\u201d]+["\u201d])(\s*下拉框)选择',
            r'\1\2中选择', new)
        if new2 != new:
            self._record_fix('L1', 'R18', sheet, case_name, step_num,
                             new, new2)
            new = new2

        # 18c: 统一"时间选择框选择"为"时间选择框中选择"
        new2 = re.sub(
            r'(在["\u201c][^"\u201d]+["\u201d])(\s*时间选择框)选择',
            r'\1\2中选择', new)
        if new2 != new:
            self._record_fix('L1', 'R18', sheet, case_name, step_num,
                             new, new2)
            new = new2

        # 18c-2: 统一“选择框”为“下拉框” (per convention: 选择框 = el-select)
        new2 = re.sub(r'(?<!级联)(?<!时间)选择框', '下拉框', new)
        if new2 != new:
            self._record_fix('L1', 'R18', sheet, case_name, step_num,
                             new, new2)
            new = new2

        # 18d: 去除"输入框，输入"中的多余逗号 → "输入框中输入"
        #   在"工单标题"输入框，输入"XX" → 在"工单标题"输入框中输入"XX"
        new2 = re.sub(
            r'(在["\u201c][^"\u201d]+["\u201d]\s*(?:输入框|文本框))[，,]\s*输入',
            r'\1中输入', new)
        if new2 != new:
            self._record_fix('L1', 'R18', sheet, case_name, step_num,
                             new, new2)
            new = new2

        return new

    def _r21_filter_comments(self, content: str, sheet: str,
                             case_name: str, step_num: int) -> str:
        """R21: 过滤注释续行 — 以 # 或 // 开头的续行删除"""
        lines = content.split('\n')
        filtered = [l for l in lines if not re.match(r'^\s*(?:#|//)', l.strip())]
        new_content = '\n'.join(filtered).strip()
        if new_content != content:
            self._record_fix('L1', 'R21', sheet, case_name, step_num,
                             content, new_content)
        return new_content

    def _r22_filter_separators(self, content: str, sheet: str,
                               case_name: str, step_num: int) -> str:
        """R22: 过滤分隔符行 — 纯 --- / === / *** 的行删除"""
        lines = content.split('\n')
        filtered = [l for l in lines if not re.match(r'^\s*[-=*]{3,}\s*$', l.strip())]
        new_content = '\n'.join(filtered).strip()
        if new_content != content:
            self._record_fix('L1', 'R22', sheet, case_name, step_num,
                             content, new_content)
        return new_content

    # ─── 新增 L1 规则 (v3) ───

    def _r31_inline_comment_strip(self, content: str, sheet: str,
                                   case_name: str, step_num: int) -> str:
        """R31: 内嵌注释去除 — 仅去除含特定关键词的括号注释 [v3]

        安全设计:
        - 模式1: 括号内含 替换/替代/参考/可以用/可替换/可用/随机数/例如 关键词
        - 模式2: 括号内纯 ASCII（非中文 = 不太可能是有效值）
        """
        new = content
        # 模式1: 明确的替换/参考注释
        new = re.sub(
            r'[（(]([^）)""\'"]{2,50}?'
            r'(?:替换|替代|参考|可以用|可替换|可用|随机数|例如)[^）)""\'"]*?)[）)]$',
            '', new).strip()
        # 模式2: 括号内纯英文/数字注释（仅在模式1未匹配时）
        if new == content:
            m = re.search(r'[（(]([A-Za-z0-9\s._-]{2,30})[）)]$', new)
            if m:
                new = new[:m.start()].strip()
        if new and new != content:
            self._record_fix('L1', 'R31', sheet, case_name, step_num,
                             content, new)
        return new

    def _r33_dropdown_no_prefix_fix(self, content: str, sheet: str,
                                     case_name: str, step_num: int) -> str:
        """R33: 缺少 "在" 前缀的下拉框操作 [v3]

        "XX下拉框选择"YY"" → "在"XX"下拉框中选择"YY""
        """
        new = content
        m = re.match(r'^(\S{2,10}?)下拉框选择([""""].+)$', new)
        if m:
            field = m.group(1)
            rest = m.group(2)
            if not re.match(r'^[""""]', field):
                field = f'"{field}"'
            new = f'在{field}下拉框中选择{rest}'
            self._record_fix('L1', 'R33', sheet, case_name, step_num,
                             content, new)
        return new

    def _r36_find_to_click(self, content: str, sheet: str,
                            case_name: str, step_num: int) -> str:
        """R36: "找到XX区域/部分" → "点击"XX"区域" [v3]

        语义变化: 在天枢系统上下文中 "找到XX区域" 通常指点击该区域。
        """
        new = content
        m = re.match(r'^找到(.+?)(区域|部分)$', new)
        if m:
            target = m.group(1)
            suffix = m.group(2)
            if not re.match(r'^[""""]', target):
                target = f'"{target}"'
            new = f'点击{target}{suffix}'
            self._record_fix('L1', 'R36', sheet, case_name, step_num,
                             content, new)
        return new

    def _r32_go_back_normalize(self, content: str, sheet: str,
                                case_name: str, step_num: int) -> str:
        """R32: "返回XX页面" → "返回" [v3]

        go_back 仅匹配裸 "返回"（l3_call 最高优先级的 "返回" 是已注册关键字）。
        "返回总览页面" 会被 l3_call 误匹配为非注册关键字。
        """
        new = content
        m = re.match(r'^返回(?:到)?(.+)$', new)
        if m and m.group(1).strip():
            new = '返回'
            self._record_fix('L1', 'R32', sheet, case_name, step_num,
                             content, new)
        return new

    def _r23_assertion_format(self, content: str, sheet: str,
                               case_name: str, step_num: int) -> str:
        """R23: 断言格式统一 — "断言可见：XX" → "断言：可见"XX"" [v3]"""
        new = content
        # 模式1: "断言可见：XX" → "断言：可见"XX""
        m = re.match(r'^断言可见[：:]\s*(.+)$', new)
        if m:
            val = m.group(1).strip()
            if not re.search(r'["""\']', val):
                val = f'"{val}"'
            new = f'断言：可见{val}'
        # 模式2: "断言"后无冒号 → 补冒号（安全兜底）
        if new == content:
            m = re.match(r'^(断言)([^：:，,\s].*)$', new)
            if m and '断言可见' not in new:
                new = f'断言：{m.group(2)}'
        if new != content:
            self._record_fix('L1', 'R23', sheet, case_name, step_num,
                             content, new)
        return new

    def _r24_custom_keyword_convert(self, content: str, sheet: str,
                                     case_name: str, step_num: int) -> str:
        """R24: 自定义断言关键词 → 标准 assert 格式 [v3]

        [v3-audit 重构] 通用逻辑，不硬编码任何 L3 关键字名称:
        1. 如果文本匹配 l3_call 格式（中文关键字+可选参数）且关键字已注册
           → 保持原样，让 l3_call + L3 管线处理（workflow 可展开为多步操作）
        2. 仅对未注册的文本执行 assert 降级转换（特定模式匹配）
        """
        new = content

        # ── 通用 L3 保护: 已注册 workflow 关键字一律保留 ──
        try:
            wf_names = set(self._get_workflow_cache().keys())
        except Exception:
            wf_names = set()

        if wf_names:
            # 提取 l3_call 格式的关键字（与 step_patterns.py 同模式）
            m_l3 = re.match(r'^([一-鿿A-Za-z_]\w{1,7})\s*(?:[（(](.+?)[）)])?\s*$', new)
            if m_l3 and m_l3.group(1) in wf_names:
                # 已注册的 L3 workflow → 不转换，保持 l3_call
                return content

        # ── 以下仅处理未注册/非 l3_call 格式的文本 ──

        # "首页列表校验(XX)" → "断言：可见首页"XX"列表"
        m = re.match(r'^首页列表校验[（(](.+?)[）)]$', new)
        if m:
            new = f'断言：可见首页"{m.group(1)}"列表'

        # "检查站内信显示(XX)" → "断言：可见站内信"XX""
        if new == content:
            m = re.match(r'^检查站内信显示[（(](.+?)[）)]$', new)
            if m:
                new = f'断言：可见站内信"{m.group(1)}"'

        # "确认下方XX显示" → "断言：可见下方"XX""
        if new == content:
            m = re.match(r'^确认下方(.+?)显示$', new)
            if m:
                new = f'断言：可见下方"{m.group(1).strip()}"'

        # "验证显示XX" → "断言：可见"XX""
        if new == content:
            m = re.match(r'^验证显示(.+)$', new)
            if m and len(m.group(1)) < 30:
                new = f'断言：可见"{m.group(1).strip()}"'

        if new != content:
            self._record_fix('L1', 'R24', sheet, case_name, step_num,
                             content, new)
        return new

    def _r26_input_field_normalize(self, content: str, sheet: str,
                                    case_name: str, step_num: int) -> str:
        """R26: 输入框格式补齐 [v3]

        - "XX"框中输入 → "XX"输入框中输入
        - "XX"框输入 → "XX"输入框中输入
        - 在"XX"中输入 → 在"XX"输入框中输入
        """
        new = content
        # 模式1: "XX"框中输入 → "XX"输入框中输入
        new = re.sub(
            r'(在[""""][^""""]+["""""])'
            r'(?<![输入下拉选择文本])框中输入',
            r'\1输入框中输入', new)
        # 模式2: "XX"框输入 → "XX"输入框中输入
        if new == content:
            new = re.sub(
                r'(在[""""][^""""]+["""""])'
                r'(?<![输入下拉选择文本])框输入',
                r'\1输入框中输入', new)
        # 模式3: 在"XX"中输入 → 在"XX"输入框中输入
        m = re.match(
            r'^(在[""""]([^""""]+)["""""])中输入([""""].+)$', new)
        if m and not any(kw in m.group(2)
                         for kw in ['输入框', '文本框', '框', '下拉框', '选择框',
                                    '级联框', '时间选择框']):
            new = m.group(1) + '输入框中输入' + m.group(3)
        if new != content:
            self._record_fix('L1', 'R26', sheet, case_name, step_num,
                             content, new)
        return new

    def _r27_wait_normalize(self, content: str, sheet: str,
                             case_name: str, step_num: int) -> str:
        """R27 v3: 等待格式标准化（仅模式1+4）

        v3 变更: 删除模式2（回归bug）和模式3（不必要）。
        - 模式1: "等待进入XX页面" → "等待XX页面加载完成"
        - 模式4: "等待XX，至少/确保/并且/同时" → 剥离尾部（不含"断言"）
        """
        new = content
        # 模式1: "等待进入XX页面" → "等待XX页面加载完成"
        m = re.match(r'^等待进入(.+页面)$', new)
        if m:
            new = f'等待{m.group(1)}加载完成'
        # 模式4: 等待XX，至少/并且/确保/同时 → 剥离尾部
        if new == content:
            m = re.match(
                r'^(等待[^，,]+)[，,]\s*(?:至少|并且|确保|同时)', new)
            if m:
                new = m.group(1).strip()
        if new != content:
            self._record_fix('L1', 'R27', sheet, case_name, step_num,
                             content, new)
        return new

    def _r28_navigate_normalize(self, content: str, sheet: str,
                                 case_name: str, step_num: int) -> str:
        """R28: 导航/跳转 — 点击XX可以跳转 → 点击"XX"可以跳转 [v3]"""
        new = content
        m = re.match(r'^点击([^""""]{1,20})可以跳转', new)
        if m:
            new = f'点击"{m.group(1)}"可以跳转'
            self._record_fix('L1', 'R28', sheet, case_name, step_num,
                             content, new)
        return new

    def _r29_dropdown_fix(self, content: str, sheet: str,
                           case_name: str, step_num: int) -> str:
        """R29: 下拉框操作修复 [v3]

        - 模式1: "XX下拉框中输入" → "XX下拉框中选择"
        - 模式2: 下拉框选择值缺引号 → 补引号
        """
        new = content
        # 模式1: "XX下拉框中输入" → "XX下拉框中选择"
        if re.search(r'["""""]下拉框中输入', new):
            new = new.replace('下拉框中输入', '下拉框中选择', 1)
        # 模式2: 下拉框选择值缺引号 → 补引号
        m = re.match(
            r'^(在[""""][^""""]+["""""]下拉框中选择)'
            r'([^""""]{1,}[^，,]*?)((?:[，,]等待.+)?)$', new)
        if m:
            val = m.group(2).strip()
            if val and not re.match(r'^[""""]', val):
                new = f'{m.group(1)}"{val}"{m.group(3)}'
        if new != content:
            self._record_fix('L1', 'R29', sheet, case_name, step_num,
                             content, new)
        return new

    def _r25_split_wait_assertion(self, text: str, sheet: str,
                                   case_name: str) -> str:
        """R25: 等待+断言/条件 复合拆分 [v3]

        在 _apply_l1_rules 的 rebuilt 阶段运行。
        与 R20 互不冲突: R25 处理 等待+断言, R20 处理 操作+等待Ns。
        """
        steps = self._parse_steps(text)
        if not steps:
            return text

        new_steps = []
        split_found = False

        for step_num, content in steps:
            m = re.match(
                r'^(等待[^，,]+)[，,]\s*(断言|至少|确保|并且|同时)(.*)$', content)
            if m:
                wait_part = m.group(1).strip()
                cond_keyword = m.group(2)
                cond_rest = m.group(3).strip()

                new_steps.append((step_num, wait_part))
                if cond_keyword == '断言':
                    desc = cond_rest if cond_rest else '操作成功'
                    new_steps.append((step_num, f'断言：{desc}'))
                else:
                    desc = f'{cond_keyword}{cond_rest}'
                    new_steps.append((step_num, f'断言：{desc}'))
                split_found = True
                continue

            new_steps.append((step_num, content))

        if not split_found:
            return text

        renumbered = [(i, s[1]) for i, s in enumerate(new_steps, 1)]
        result = self._rebuild_steps(renumbered)
        self._record_fix('L1', 'R25', sheet, case_name, 0,
                         '存在等待+断言复合步骤', '已拆分')
        return result

    def _r34_split_action_wait_condition(self, text: str, sheet: str,
                                          case_name: str) -> str:
        """R34: 操作 + 等待文本 + 条件 三段复合拆分 [v3]

        在 _apply_l1_rules 的 rebuilt 阶段运行。
        处理 R20（操作+等待Ns）和 R25（等待+断言）都不覆盖的三段模式。
        例: 点击"查询"按钮，等待查询结果展示成功，至少有一条问题记录
        """
        from core.step_patterns import parse_step

        steps = self._parse_steps(text)
        if not steps:
            return text

        new_steps = []
        split_found = False

        for step_num, content in steps:
            m = re.match(
                r'^(.+?)[，,]\s*(等待[^，,]+)[，,]\s*'
                r'(断言|至少|确保|并且|同时)(.*)$', content)
            if m:
                action_part = m.group(1).strip()
                wait_part = m.group(2).strip()
                cond_keyword = m.group(3)
                cond_rest = m.group(4).strip()

                # 仅当操作部分是有效的 click/fill/el_select 等才拆分
                parsed = parse_step(action_part)
                if parsed['type'] != 'unknown':
                    new_steps.append((step_num, action_part))
                    new_steps.append((step_num, wait_part))
                    if cond_keyword == '断言':
                        desc = cond_rest if cond_rest else '操作成功'
                        new_steps.append((step_num, f'断言：{desc}'))
                    else:
                        new_steps.append((step_num, f'断言：{cond_keyword}{cond_rest}'))
                    split_found = True
                    continue

            new_steps.append((step_num, content))

        if not split_found:
            return text

        renumbered = [(i, s[1]) for i, s in enumerate(new_steps, 1)]
        result = self._rebuild_steps(renumbered)
        self._record_fix('L1', 'R34', sheet, case_name, 0,
                         '存在操作+等待+条件复合步骤', '已拆分')
        return result

    def _r04_split_merged(self, text: str, sheet: str, case_name: str) -> str:
        """R04: 合并步骤拆分
        只匹配中文字符后紧跟的步骤编号（避免 URL 中的 100.71 或版本号 2.1.3 被误识别）
        """
        lines = text.split('\n')
        new_lines = []
        merged_found = False

        # 匹配：中文字符（排除标点）+ 数字 + 步骤分隔符
        # 例: "展示成功4. 选择第一条" → 匹配
        # 不匹配: "http://100.71" 或 "（2.1.3）" 或 "版本3. 说明"
        merged_re = re.compile(r'([一-鿿㐀-䶿])\s*(\d+)\s*[.、．]\s*(.+)')

        for line in lines:
            # 解析当前步骤
            m = re.match(r'^(\d+\s*[.、．]\s*)(.+)$', line, re.DOTALL)
            if not m:
                new_lines.append(line)
                continue

            prefix = m.group(1)
            content = m.group(2)

            # 检查内容中是否嵌入另一个步骤编号（前面必须有中文字符）
            inner = merged_re.search(content)
            if inner:
                part1 = content[:inner.start() + len(inner.group(1))].strip()
                part2 = f"{inner.group(2)}. {inner.group(3).strip()}"
                new_lines.append(f"{prefix}{part1}")
                new_lines.append(part2)
                merged_found = True
            else:
                new_lines.append(line)

        if merged_found:
            result = '\n'.join(new_lines)
            steps = self._parse_steps(result)
            if steps:
                result = self._rebuild_steps([(i, s[1]) for i, s in
                                              enumerate(steps, 1)])
                self._record_fix('L1', 'R04', sheet, case_name, 0,
                                 '存在合并步骤', '已拆分为独立步骤')
                return result

        return text

    def _r06_split_assertions(self, text: str, sheet: str, case_name: str) -> str:
        """R06: 断言拆分 — 断言：A、B、C可见 → 多行"""
        lines = text.split('\n')
        new_lines = []
        split_found = False

        for line in lines:
            m = re.match(r'^(\d+\s*[.、．])\s*断言[：:](.+?)[、，,](.+?)可见\s*$', line)
            if m:
                prefix = m.group(1)
                items_str = m.group(0)
                # 提取 "断言：XXX、YYY、ZZZ可见" 中的所有项
                assertion_match = re.search(r'断言[：:](.+?)可见', line)
                if assertion_match:
                    items_text = assertion_match.group(1)
                    items = re.split(r'[、，,]', items_text)
                    if len(items) > 1:
                        # 拆分为多行
                        for item in items:
                            item = item.strip()
                            if item:
                                new_lines.append(f"{prefix} 断言：{item}可见")
                                prefix = self._next_step_prefix(prefix)
                        split_found = True
                        continue

            new_lines.append(line)

        if split_found:
            result = '\n'.join(new_lines)
            steps = self._parse_steps(result)
            if steps:
                result = self._rebuild_steps([(i, s[1]) for i, s in
                                              enumerate(steps, 1)])
                self._record_fix('L1', 'R06', sheet, case_name, 0,
                                 '多目标断言合并', '拆分为独立断言行')
                return result

        return text

    def _r20_split_compound_steps(self, text: str, sheet: str,
                                   case_name: str) -> str:
        """R20: 复合步骤拆分 — 操作...，等待Ns → 两个独立步骤

        匹配步骤末尾的 "，等待Ns" 或 ",等待Ns" 后缀，
        将其拆分为操作步骤 + 等待步骤。仅当操作部分非空时拆分。

        例:
          原始: 在"方案名称"下拉框中选择"移动云资源池22期工程方案三"，等待1s
          拆为: ① 在"方案名称"下拉框中选择"移动云资源池22期工程方案三"
                ② 等待1s
        """
        steps = self._parse_steps(text)
        if not steps:
            return text

        new_steps = []
        split_found = False

        # 匹配: 操作内容 + 逗号 + 等待（在步骤末尾）
        # 支持两种等待格式: ① 等待Ns  ② L3 等待关键字（等待加载完成/等待页面加载完成）
        _WAIT_SUFFIX = r'(?:等待\s*\d+\s*s|等待加载完成|等待页面加载完成)'
        compound_re = re.compile(
            rf'^(.+?)[，,]\s*({_WAIT_SUFFIX})\s*$')

        for step_num, content in steps:
            m = compound_re.match(content)
            if m:
                action_part = m.group(1).strip()
                wait_part = m.group(2).strip()
                # 仅当操作部分非空且不是纯"等待"时才拆分
                if action_part and not re.match(
                    rf'^{_WAIT_SUFFIX}$', action_part):
                    # el-select / 级联选择器：R11/R13 已正确追加等待，不拆分
                    if re.match(r'^在"[^"]+"(?:下拉框|级联选择框)中选择"[^"]+"$', action_part):
                        new_steps.append((step_num, content))  # 保留整体
                    else:
                        new_steps.append((step_num, action_part))
                        new_steps.append((step_num, wait_part))
                        split_found = True
                    continue

            new_steps.append((step_num, content))

        if not split_found:
            return text

        # 重新编号
        renumbered = [(i, s[1]) for i, s in enumerate(new_steps, 1)]
        result = self._rebuild_steps(renumbered)

        self._record_fix('L1', 'R20', sheet, case_name, 0,
                         '存在复合步骤（操作+等待合并）',
                         '已拆分为独立步骤')
        return result

    def _dedup_consecutive_waits(self, text: str, sheet: str, case_name: str) -> str:
        """R20b: 去重连续重复的等待步骤

        覆盖两种重复模式:
          A) 两步完全相同: 等待1s → 等待1s
          B) 前一步已含等待后缀 + 下一步是纯等待:
             在"XX"下拉框中选择"YY"，等待1s → 等待1s
        """
        steps = self._parse_steps(text)
        if not steps:
            return text

        deduped = []
        dup_count = 0
        _WAIT_PATTERN = r'(?:等待\s*\d+\s*s|等待加载完成|等待页面加载完成)'
        wait_re = re.compile(rf'^{_WAIT_PATTERN}$')
        trailing_wait_re = re.compile(rf'[，,]\s*{_WAIT_PATTERN}\s*$')

        for step_num, content in steps:
            if deduped and wait_re.match(content):
                prev_content = deduped[-1][1]
                # 模式 A: 完全相同
                if prev_content == content:
                    dup_count += 1
                    continue
                # 模式 B: 前一步已含等待后缀
                if trailing_wait_re.search(prev_content):
                    dup_count += 1
                    continue
            deduped.append((step_num, content))

        if dup_count == 0:
            return text

        renumbered = [(i, s[1]) for i, s in enumerate(deduped, 1)]
        result = self._rebuild_steps(renumbered)
        self._record_fix('L1', 'R20b', sheet, case_name, 0,
                         f'存在 {dup_count} 处连续重复等待步骤',
                         '已去重')
        return result

    # ─── R37: 引号完整性（补充 _enforce_quotes 的缺口）───

    def _r37_quote_integrity(self, text: str, sheet: str, case_name: str) -> str:
        """R37: 引号完整性 — 确保已分类步骤的操作目标有引号。

        _enforce_quotes 只在 parse_step → unknown 时触发。
        本规则覆盖已被 parse_step 正确分类但操作目标缺少引号的步骤。
        """
        from core.step_patterns import parse_step

        steps = self._parse_steps(text)
        if not steps:
            return text

        modified = False
        new_steps = []

        for step_num, content in steps:
            original = content

            # Skip if already has quotes
            if re.search(r'[""""「]', content):
                new_steps.append((step_num, content))
                continue

            # Skip unknown steps (handled by _enforce_quotes)
            parsed = parse_step(content)
            if parsed.get('type') == 'unknown':
                new_steps.append((step_num, content))
                continue

            # Pattern 1: 点击XX按钮/链接/标签页 → 点击"XX"按钮/链接/标签页
            m = re.match(
                r'^(点击)([一-鿿A-Za-z0-9]{2,8})(按钮|链接|标签页|复选框|开关|图标)$',
                content)
            if m:
                action, target, suffix = m.group(1), m.group(2), m.group(3)
                content = f'{action}"{target}"{suffix}'

            # Pattern 4: 点击XX跳转/打开/展开/查看/切换/筛选/搜索 → 点击"XX"跳转...
            # 必须在 Pattern 2 之前运行，否则 Pattern 2 贪心匹配把 "里程碑跳转" 整体当作 target
            # 解决 Phase 6 label 提取失败: 无引号 → label='' → KB/discovery/fallback 全部跳过
            if content == original:
                m = re.match(
                    r'^(点击)([一-鿿A-Za-z0-9]{2,8})'
                    r'(跳转|打开|展开|查看|切换|筛选|搜索)$',
                    content)
                if m:
                    action, target, verb = m.group(1), m.group(2), m.group(3)
                    content = f'{action}"{target}"{verb}'

            # Pattern 2: 点击XX (2-8 char pure Chinese/English target, no suffix)
            if content == original:
                m = re.match(r'^(点击)([一-鿿A-Za-z]{2,8})$', content)
                if m:
                    action, target = m.group(1), m.group(2)
                    content = f'{action}"{target}"'

            # Pattern 3: 勾选/取消勾选XX → 勾选"XX"
            # 排除 check_first 模式（勾选第N个/条XX），该模式不支持引号
            if content == original:
                # check_first 模式: 勾选第N个/条XX — 不加引号
                if re.match(r'^勾选第[一二三四五六七八九十\d]+[个条]?', content):
                    pass  # 跳过 R37，保持原样
                else:
                    m = re.match(r'^(勾选|取消勾选|选中)([一-鿿A-Za-z]{2,8})$', content)
                    if m:
                        action, target = m.group(1), m.group(2)
                        content = f'{action}"{target}"'

            if content != original:
                modified = True
                self._record_fix('L1', 'R37', sheet, case_name, step_num,
                                 original, content)

            new_steps.append((step_num, content))

        if not modified:
            return text
        return self._rebuild_steps(new_steps)

    # ─── 引号强制执行 (L1 最后一步) ───

    def _enforce_quotes(self, text: str, sheet: str, case_name: str) -> str:
        """对每条步骤执行: parse_step → unknown → 尝试加引号修复 → 重验证 [v3]"""
        from core.step_patterns import parse_step

        steps = self._parse_steps(text)
        if not steps:
            return text

        new_steps = []

        for step_num, content in steps:
            parsed = parse_step(content)
            if parsed['type'] != 'unknown':
                new_steps.append((step_num, content))
                continue

            # 尝试结构性引号补全
            fixed = self._try_quote_fix(content)
            if fixed and fixed != content:
                reparsed = parse_step(fixed)
                if reparsed['type'] != 'unknown':
                    self._record_fix('L1', 'RQ', sheet, case_name, step_num,
                                     content, fixed)
                    new_steps.append((step_num, fixed))
                    continue

            new_steps.append((step_num, content))

        return self._rebuild_steps(new_steps)

    def _try_quote_fix(self, text: str):
        """基于结构性关键词尝试给未引号文本加引号 [v3]

        仅在 parse_step() 返回 unknown 时调用（已有类型的步骤不受影响）。
        返回修复后的文本，或 None 表示无法修复。
        """
        # 模式1: 在XX下拉框中选择/输入 → 在"XX"下拉框中选择/输入
        m = re.match(
            r'^在(\S{2,10}?)'
            r'((?:下拉框|输入框|文本框|时间选择框|级联选择框)'
            r'中(?:选择|输入).+)$', text)
        if m and not re.match(r'^[""""]', m.group(1)):
            return f'在"{m.group(1)}"{m.group(2)}'

        # 模式2: 点击第N条记录的XX按钮 → 加引号（比模式3更具体，必须先匹配）
        m = re.match(r'^(点击第.+?记录的)([^\s""""]{1,15})(按钮.*)$', text)
        if m:
            return f'{m.group(1)}"{m.group(2)}"{m.group(3)}'

        # 模式3: 点击XX按钮 → 点击"XX"按钮
        m = re.match(r'^点击([^\s""""]{1,15})按钮', text)
        if m:
            return f'点击"{m.group(1)}"按钮'

        # 模式4b: 点击XX跳转/打开/展开/查看/切换/筛选/搜索 → 点击"XX"跳转...
        # 必须在模式4之前：模式4 的 $ 锚点会贪心匹配 "点击详情跳转" → 点击"详情跳转"（错误）
        # 覆盖 R37 Pattern 4 无法处理的 unknown 类型步骤（target 较长时 parse_step 归为 unknown）
        m = re.match(
            r'^(点击)([一-鿿A-Za-z0-9]{2,8})'
            r'(跳转|打开|展开|查看|切换|筛选|搜索)$', text)
        if m:
            return f'{m.group(1)}"{m.group(2)}"{m.group(3)}'

        # 模式4: 点击XX → 点击"XX"（短文本，2-6字符，纯中文/英文，无后缀）
        m = re.match(r'^点击([一-鿿A-Za-z_]{2,6})$', text)
        if m:
            return f'点击"{m.group(1)}"'

        # 模式5: 在XX中输入/里输入"YY" → 在"XX"中输入"YY"
        m = re.match(
            r'^在(\S{2,10}?)(中|里)输入([""""].+)$', text)
        if m and not re.match(r'^[""""]', m.group(1)):
            return f'在"{m.group(1)}"{m.group(2)}输入{m.group(3)}'

        return None

    def _next_step_prefix(self, prefix: str) -> str:
        """计算下一步的编号前缀"""
        m = re.match(r'(\d+)', prefix)
        if m:
            return f"{int(m.group(1)) + 1}."
        return prefix

    # ─── L2 规则（自动修复 + 标记） ───

    def _apply_l2_rules(self, text: str, sheet: str, case_name: str,
                        row_idx: int) -> str:
        """应用 L2 规则"""
        steps = self._parse_steps(text)
        if not steps:
            return text

        new_steps = []
        for step_num, content in steps:
            original = content

            # R11: el-select 缺等待
            content = self._r11_select_wait(content, sheet, case_name, step_num)

            # R12: 组件类型-时间
            content = self._r12_time_component(content, sheet, case_name, step_num)

            # R13: 级联选择器
            content = self._r13_cascader(content, sheet, case_name, step_num)

            # R14: 弹窗描述匹配
            content = self._r14_dialog_desc(content, sheet, case_name, step_num)

            # R15: 值校验断言格式
            content = self._r15_value_assertion(content, sheet, case_name, step_num)

            # R16: 操作对象前缀（仅标记）
            self._r16_action_target(content, sheet, case_name, step_num, steps)

            # R17: "更多"菜单前缀（仅标记）
            self._r17_more_menu(content, sheet, case_name, step_num, steps)

            # M17: R37 断言语义规范化（在 R23/R24 之后执行）
            content = self._r37_assertion_normalize(content, sheet, case_name, step_num)

            new_steps.append((step_num, content))

        return self._rebuild_steps(new_steps)

    def _r11_select_wait(self, content: str, sheet: str,
                         case_name: str, step_num: int) -> str:
        """R11: el-select 缺等待 — 补充 "等待1s" """
        # 匹配: 在"XX"下拉框中选择"YY" （但后面没有 "等待"）
        m = re.match(r'^(在"[^"]+"(?:下拉框|级联选择框)中选择"[^"]+")\s*$', content)
        if m:
            new = f"{m.group(1)}，等待1s"
            self._record_fix('L2', 'R11', sheet, case_name, step_num,
                             content, new)
            return new
        return content

    def _r12_time_component(self, content: str, sheet: str,
                            case_name: str, step_num: int) -> str:
        """R12: 时间相关字段的"下拉框" → "时间选择框" """
        time_keywords = ['时间', '日期', '发函', '开始', '结束', '投诉时间',
                         '发生时间']
        for kw in time_keywords:
            if kw in content:
                m = re.search(rf'在"([^"]*{re.escape(kw)}[^"]*)"下拉框中选择', content)
                if m:
                    field = m.group(1)
                    new = content.replace(
                        f'在"{field}"下拉框中选择',
                        f'在"{field}"时间选择框中选择')
                    self._record_fix('L2', 'R12', sheet, case_name, step_num,
                                     content, new)
                    return new
        return content

    def _r13_cascader(self, content: str, sheet: str,
                      case_name: str, step_num: int) -> str:
        """R13: 级联选择器识别 — "点击XX，点击A，再点击B" → 标准格式"""
        # 匹配: 点击"XX"输入框，点击"A"，再点击"B"
        m = re.match(
            r'^点击"([^"]+)"输入框，点击"([^"]+)"，再点击"([^"]+)"(.*)$',
            content)
        if m:
            field, level1, level2, rest = m.groups()
            new = f'在"{field}"级联选择框中依次选择"{level1}"、"{level2}"'
            if '等待' not in rest:
                new += '，等待1s'
            elif rest.strip().startswith('，'):
                new += rest.strip()
            self._record_fix('L2', 'R13', sheet, case_name, step_num,
                             content, new)
            return new
        return content

    def _r14_dialog_desc(self, content: str, sheet: str,
                         case_name: str, step_num: int) -> str:
        """R14: 弹窗描述匹配 — "确认删除的弹窗" 出现在非删除操作中"""
        # 检测：步骤中提到"确认删除"但用例名不含"删除"
        if '删除' not in case_name and '确认删除的弹窗' in content:
            # 尝试从用例名推断正确的弹窗描述
            for action in ['阻塞报备', '归档', '投诉', '编辑', '新增', '修复']:
                if action in case_name:
                    new = content.replace('确认删除的弹窗', f'{action}弹窗')
                    self._record_fix('L2', 'R14', sheet, case_name, step_num,
                                     content, new)
                    return new
        return content

    def _r15_value_assertion(self, content: str, sheet: str,
                             case_name: str, step_num: int) -> str:
        """R15: 值校验断言 — "检查/查看XX为'YY'" → "断言：XX为'YY'可见" """
        patterns = [
            # "检查列表中第一条记录的XX为"YY""
            (r'^(检查|查看)(列表中第一条记录的.+?)为["\u201c](.+?)["\u201d]\s*$',
             lambda m: f'断言：{m.group(2)}为"{m.group(3)}"可见'),
            # "返回XX后，查看第一条记录的XX为"YY""
            (r'^返回(.+?)后，(查看|检查)(.+?)为["\u201c](.+?)["\u201d]\s*$',
             lambda m: f'断言：返回{m.group(1)}后，{m.group(3)}为"{m.group(4)}"可见'),
        ]
        for pat, transform in patterns:
            m = re.match(pat, content)
            if m:
                new = transform(m)
                self._record_fix('L2', 'R15', sheet, case_name, step_num,
                                 content, new)
                return new
        return content

    def _r16_action_target(self, content: str, sheet: str,
                           case_name: str, step_num: int,
                           all_steps: List[Tuple[int, str]]):
        """R16: 操作对象前缀 — 仅标记"""
        # 检测: 直接 "点击编辑/删除/修复方案" 但前面没有行选择
        action_buttons = ['编辑', '删除', '修复方案', '导出', '进展更新',
                          'PMO更新', '保存']
        for btn in action_buttons:
            if re.match(rf'^点击["\u201c]{btn}["\u201d]\s*$', content):
                # 检查前一步是否有行选择
                prev_has_row_select = False
                for sn, sc in all_steps:
                    if sn == step_num - 1:
                        if any(kw in sc for kw in ['第一条', '查询记录', '选择第']):
                            prev_has_row_select = True
                        break

                if not prev_has_row_select:
                    self._record_fix('L2', 'R16', sheet, case_name, step_num,
                                     content,
                                     f'建议补充: 选择第一条查询记录，{content}',
                                     auto_fixed=False)

    def _r17_more_menu(self, content: str, sheet: str,
                       case_name: str, step_num: int,
                       all_steps: List[Tuple[int, str]]):
        """R17: "更多"菜单前缀 — 仅标记"""
        if re.match(r'^点击["\u201c]更多["\u201d]', content):
            prev_has_row_select = False
            for sn, sc in all_steps:
                if sn == step_num - 1:
                    if any(kw in sc for kw in ['第一条', '查询记录', '选择第']):
                        prev_has_row_select = True
                    break

            if not prev_has_row_select:
                self._record_fix('L2', 'R17', sheet, case_name, step_num,
                                 content,
                                 f'建议补充: 选择第一条查询记录，{content}',
                                 auto_fixed=False)

    def _r37_assertion_normalize(self, content: str, sheet: str,
                                  case_name: str, step_num: int) -> str:
        """M17: R37 断言语义规范化 — 消除模糊断言的 NLP 解析歧义

        模式1: "等待X展示成功" → "等待X加载完成" (消除"成功"对下游的误导)
        模式3: "断言：返回X后可见" → "断言：可见"X"" (保留核心名词)
        注: 模式2（数量断言降级为可见性）已删除，由 assert_count 处理
        """
        new = content

        # 模式1: "等待X展示成功" → "等待X加载完成"
        m = re.match(r'^等待(.+?)展示成功$', new)
        if m:
            new = f'等待{m.group(1)}加载完成'

        # 模式3: "断言：返回X后可见" → "断言：可见"X""
        if new == content:
            m = re.match(r'^断言[：:]\s*返回(.+?)后可见$', new)
            if m:
                val = m.group(1).strip()
                new = f'断言：可见"{val}"'

        if new != content:
            self._record_fix('L2', 'R37', sheet, case_name, step_num,
                             content, new)

        return new

    # ─── L3: 解析能力验证 ───

    def _apply_l3_parse_validation(self, text: str, sheet: str, case_name: str):
        """L3a: 验证每个步骤都能被 parse_step() 分类

        unknown type = error（阻断）。为每个 unknown 步骤生成修改建议。
        """
        from core.step_patterns import parse_step  # 强制依赖，缺失即报错

        steps = self._parse_steps(text)
        for step_num, content in steps:
            parsed = parse_step(content)
            if parsed['type'] == 'unknown':
                suggestion = self._generate_suggestion(content)
                self._record_fix(
                    'L3', 'R19', sheet, case_name, step_num,
                    content,
                    '此步骤无法被解析系统识别',
                    auto_fixed=False,
                    severity='error',
                    suggestion=suggestion,
                )

    # ─── L3: 关键字白名单 + 参数校验 ───

    def _apply_l3_whitelist_validation(self, text: str, sheet: str, case_name: str):
        """L3b/c: l3_call 白名单检查 + 参数数量验证

        - workflow 未找到 → 先尝试回退重解析（与 generate_step 一致）
          - 回退成功 → 跳过（不是真正的 L3 调用）
          - 回退也失败 → error + 列出可用 workflow
        - 参数不足 → error
        """
        from core.step_patterns import parse_step, STEP_PATTERNS

        wf_cache = self._get_workflow_cache()
        wf_names = set(wf_cache.keys())

        steps = self._parse_steps(text)
        for step_num, content in steps:
            parsed = parse_step(content)
            if parsed['type'] != 'l3_call':
                continue

            cn_name = parsed['args'][0]
            raw_params = parsed['args'][1] or '' if len(parsed['args']) > 1 else ''

            # 白名单检查: workflow 必须存在
            if cn_name not in wf_names:
                # 回退重解析: 尝试用其他 pattern 匹配（与 generate_step 一致）
                raw_text = parsed.get('raw', '').strip()
                reparse_ok = False
                for pattern, action_type, group_names in STEP_PATTERNS:
                    if action_type == 'l3_call':
                        continue
                    m = pattern.search(raw_text)
                    if m:
                        reparse_ok = True
                        break

                if reparse_ok:
                    # 不是真正的 L3 调用，回退后能正确分类 → 跳过白名单检查
                    continue

                # 回退也失败 → 真正的 L3 错误
                available = ', '.join(sorted(wf_names))
                self._record_fix(
                    'L3', 'R20', sheet, case_name, step_num,
                    content,
                    f"'{cn_name}' 不是已注册的 L3 关键字",
                    auto_fixed=False,
                    severity='error',
                    suggestion=f"可用的 L3 关键字: {available}\n"
                               f"建议: 将操作展开为标准步骤，或在 lib/_knowledge/ 中注册新 workflow。",
                )
                continue

            # 参数数量检查
            wf_def = wf_cache[cn_name]
            param_names = wf_def.get('params', [])
            param_values = [v.strip() for v in re.split(r'[,，]', raw_params)] if raw_params else []

            if len(param_values) < len(param_names):
                missing = param_names[len(param_values):]
                self._record_fix(
                    'L3', 'R20', sheet, case_name, step_num,
                    content,
                    f"L3 关键字 '{cn_name}' 缺少参数: {', '.join(missing)}",
                    auto_fixed=False,
                    severity='error',
                    suggestion=f"请补充参数。用法: {cn_name}({', '.join(param_names)})",
                )

    def _generate_suggestion(self, step_text: str) -> str:
        """为无法解析的步骤生成修改建议

        通过关键词集合交集计算步骤文本与 42+ 模式的相似度，
        返回 top-3 最接近的标准格式示例。
        """
        from core.step_patterns import STEP_PATTERNS

        # 提取步骤文本中的中文关键词集合
        step_chars = set(re.findall(r'[一-鿿]+', step_text))

        best_matches = []  # [(score, action_type, pattern_str)]

        for pattern, action_type, group_names in STEP_PATTERNS:
            if action_type == 'l3_call':
                continue  # 跳过 l3_call（正则过于宽泛）

            # 提取正则中的中文字面量
            pattern_literals = set(re.findall(r'[一-鿿]+', pattern.pattern))
            if not pattern_literals:
                continue

            # Jaccard 相似度
            overlap = len(step_chars & pattern_literals)
            union_size = len(step_chars | pattern_literals)
            score = overlap / union_size if union_size > 0 else 0

            if score > 0.1:
                best_matches.append((score, action_type, pattern.pattern))

        best_matches.sort(key=lambda x: x[0], reverse=True)
        top = best_matches[:3]

        if not top:
            return ("当前系统不支持该操作的自动描述。\n"
                    "建议拆分为标准步骤（点击/输入/选择/断言等），"
                    "或参考 SKILL.md 中的步骤格式规范。")

        # action_type → 标准格式示例
        type_examples = {
            'click_btn': '点击"XX"按钮',
            'fill': '在"XX"输入框中输入"YY"',
            'el_select': '在"XX"下拉框中选择"YY"',
            'el_cascader': '在"XX"级联选择框中依次选择"YY"、"ZZ"',
            'click_tab': '点击"XX" tab',
            'date_select': '在"XX"时间选择框中选择"YY"',
            'dialog_date_select': '在XX弹窗中，时间选择"YY"',
            'assert': '断言：XX可见',
            'confirm_dialog': '确认"XX"',
            'confirm_delete': '确认删除...点击"XX"',
            'wait': '等待XX加载完成',
            'wait_element': '等待出现XX',
            'wait_time': '等待5s',
            'go_back': '返回',
            'refresh': '刷新',
            'open_url': '访问 https://...',
            'click_detail_link': '点击第一条记录的XX标题"TEXT"',
            'click_table_row_btn': '点击第一条记录的"XX"按钮',
            'click_first_in_list': '点击XX第一个"YY"按钮',
            'click_navigate': '点击"XX"可以跳转',
            'click_section': '点击XX部分/区域',
            'click_more_then': '点击"更多"...选择XX',
            'conditional_click_btn': '如果"XX"中数量大于0则点击"YY"按钮',
            'conditional_click_tab': '如果"XX"中数量大于0则点击"YY"tab',
            'conditional_click_row': '如果"XX"中数量大于0则点击第N条',
            'assert_row': '检查第一条状态为"XX"',
            'check_assert': '检查XX与YY一致',
            'check_first': '勾选第一个产品',
            'click_table_action': '选择第一条...，点击"XX"',
            'skip': '请完成测试结论...',
        }

        lines = ["最接近的标准格式:"]
        for score, atype, _ in top:
            example = type_examples.get(atype, f'[{atype}]')
            lines.append(f"  - {example}  (相似度 {int(score * 100)}%)")

        return '\n'.join(lines)

    def _get_workflow_cache(self):
        """加载三层 workflow 定义（系统级 + 技能级 + 项目级），缓存复用

        加载优先级: 系统级 → 技能级 → 项目级（last-writer-wins）
        """
        if not hasattr(self, '_wf_cache') or self._wf_cache is None:
            self._wf_cache = {}
            import glob as _glob
            import yaml as _yaml

            # 尝试导入 yaml
            try:
                import yaml
            except ImportError:
                return self._wf_cache

            # 推断 skill_dir: validate_excel.py 在 tools/excel/ 下
            excel_dir = os.path.dirname(os.path.abspath(__file__))
            tools_dir = os.path.dirname(excel_dir)
            skill_dir = os.path.dirname(tools_dir)

            # 层 1: 系统级
            sys_path = os.path.join(skill_dir, 'lib', 'system_workflows.yaml')
            if os.path.isfile(sys_path):
                try:
                    data = yaml.safe_load(open(sys_path, encoding='utf-8'))
                    for wf in (data or {}).get('workflows', []):
                        cn = wf.get('chinese_name')
                        if cn:
                            self._wf_cache[cn] = wf
                except Exception:
                    pass

            # 层 2: 技能级 _knowledge/*.yaml
            skill_knowledge_dir = os.path.join(skill_dir, 'lib', '_knowledge')
            if os.path.isdir(skill_knowledge_dir):
                for f in sorted(_glob.glob(os.path.join(skill_knowledge_dir, '*.yaml'))):
                    try:
                        data = yaml.safe_load(open(f, encoding='utf-8'))
                        wfs = (data or {}).get('workflows', [])
                        if isinstance(wfs, dict):
                            wfs = [{'name': k, **v} for k, v in wfs.items()]
                        for wf in (wfs or []):
                            cn = wf.get('chinese_name')
                            if cn:
                                self._wf_cache[cn] = wf
                    except Exception:
                        pass

            # 层 3: 项目级（从 Excel 路径推断项目目录）
            project_dir = self._infer_project_dir()
            if project_dir:
                knowledge_dir = os.path.join(project_dir, '_knowledge')
                if os.path.isdir(knowledge_dir):
                    for f in sorted(_glob.glob(os.path.join(knowledge_dir, '*.yaml'))):
                        if os.path.basename(f) == 'workflow_aliases.yaml':
                            continue
                        try:
                            data = yaml.safe_load(open(f, encoding='utf-8'))
                            wfs = (data or {}).get('workflows', [])
                            if isinstance(wfs, dict):
                                wfs = [{'name': k, **v} for k, v in wfs.items()]
                            for wf in (wfs or []):
                                cn = wf.get('chinese_name')
                                if cn:
                                    self._wf_cache[cn] = wf
                        except Exception:
                            pass

        return self._wf_cache

    def _infer_project_dir(self):
        """从 Excel 文件路径推断项目根目录

        常见布局:
          {project}/testcases.xlsx → project 即父目录
          {project}/docs/testcases.xlsx → project 即父父目录
        """
        excel_dir = str(self.input_path.parent)
        # 检查父目录是否有 run.py 或 _knowledge/
        if os.path.isfile(os.path.join(excel_dir, 'run.py')):
            return excel_dir
        if os.path.isdir(os.path.join(excel_dir, '_knowledge')):
            return excel_dir
        # 再往上一级
        parent = os.path.dirname(excel_dir)
        if os.path.isfile(os.path.join(parent, 'run.py')):
            return parent
        if os.path.isdir(os.path.join(parent, '_knowledge')):
            return parent
        return None

    # ─── 报告生成 ───

    def _generate_report(self):
        """生成 HTML 报告"""
        l1_auto = [f for f in self.fixes if f.level == 'L1' and f.auto_fixed]
        l2_auto = [f for f in self.fixes if f.level == 'L2' and f.auto_fixed]
        l2_warn = [f for f in self.fixes if f.level == 'L2' and not f.auto_fixed]
        l3_errors = [f for f in self.fixes if f.level == 'L3' and f.severity == 'error']

        total = len(self.fixes)

        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>Excel 预检报告</title>
<style>
body {{ font-family: -apple-system, sans-serif; max-width: 960px; margin: 40px auto;
       padding: 0 20px; color: #333; background: #f5f5f5; }}
h1 {{ color: #1a1a2e; border-bottom: 3px solid #e94560; padding-bottom: 12px; }}
.summary {{ background: #fff; border-radius: 8px; padding: 20px; margin: 20px 0;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
.summary h2 {{ margin-top: 0; }}
.stat {{ display: inline-block; padding: 8px 16px; margin: 4px; border-radius: 20px;
         font-weight: bold; font-size: 14px; }}
.stat-ok {{ background: #d4edda; color: #155724; }}
.stat-warn {{ background: #fff3cd; color: #856404; }}
.stat-err {{ background: #f8d7da; color: #721c24; }}
.stat-info {{ background: #d1ecf1; color: #0c5460; }}
.section {{ background: #fff; border-radius: 8px; margin: 16px 0;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1); overflow: hidden; }}
.section-header {{ padding: 12px 20px; font-weight: bold; font-size: 16px;
                   cursor: pointer; user-select: none; }}
.section-header:hover {{ opacity: 0.8; }}
.header-l1 {{ background: #d4edda; color: #155724; }}
.header-l2 {{ background: #cce5ff; color: #004085; }}
.header-warn {{ background: #fff3cd; color: #856404; }}
.header-l3 {{ background: #f8d7da; color: #721c24; }}
.fix-table {{ width: 100%; border-collapse: collapse; }}
.fix-table th {{ background: #f8f9fa; padding: 8px 12px; text-align: left;
                 font-size: 13px; border-bottom: 2px solid #dee2e6; }}
.fix-table td {{ padding: 8px 12px; border-bottom: 1px solid #eee;
                 font-size: 13px; vertical-align: top; }}
.fix-table tr:hover {{ background: #f8f9fa; }}
.before {{ color: #dc3545; text-decoration: line-through; }}
.after {{ color: #28a745; }}
.suggestion {{ color: #0c5460; background: #d1ecf1; padding: 4px 8px;
               border-radius: 4px; white-space: pre-wrap; }}
.rule-tag {{ display: inline-block; padding: 2px 8px; border-radius: 4px;
             font-size: 11px; font-weight: bold; background: #6c757d; color: #fff; }}
.empty-msg {{ padding: 20px; text-align: center; color: #999; }}
</style>
</head>
<body>
<h1>📋 Excel 用例预检报告</h1>
<div class="summary">
  <h2>概要</h2>
  <p>源文件: <code>{self.input_path.name}</code></p>
  <p>修正版: <code>{os.path.basename(self.output_path)}</code></p>
  <div>
    <span class="stat stat-ok">L1 自动修复: {len(l1_auto)} 处 ✓</span>
    <span class="stat stat-ok">L2 自动修复: {len(l2_auto)} 处 ✓</span>
    <span class="stat {'stat-warn' if l2_warn else 'stat-ok'}">L2 待确认: {len(l2_warn)} 处 {'⚠' if l2_warn else '✓'}</span>
    <span class="stat {'stat-err' if l3_errors else 'stat-ok'}">L3 解析错误: {len(l3_errors)} 处 {'✗' if l3_errors else '✓'}</span>
    <span class="stat stat-info">共 {total} 处</span>
  </div>
</div>
"""
        # L1 section
        html += self._render_section('L1 自动修复', 'header-l1', l1_auto)
        # L2 auto section
        html += self._render_section('L2 自动修复', 'header-l2', l2_auto)
        # L2 warn section
        html += self._render_warn_section('L2 待确认 ⚠️', l2_warn)
        # L3 error section
        html += self._render_l3_error_section('L3 解析错误 (阻断) ✗', l3_errors)

        html += """
</body>
</html>"""

        with open(self.report_path, 'w', encoding='utf-8') as f:
            f.write(html)

    def _render_section(self, title: str, header_class: str,
                        fixes: List[FixRecord]) -> str:
        if not fixes:
            return f"""
<div class="section">
  <div class="section-header {header_class}">{title} (0)</div>
  <div class="empty-msg">无此类问题</div>
</div>"""

        rows = ""
        for f in fixes:
            rows += f"""
    <tr>
      <td><span class="rule-tag">{f.rule_id}</span></td>
      <td>{f.sheet}</td>
      <td>{f.case_name}</td>
      <td>{f.step_num}</td>
      <td><span class="before">{self._esc(f.before)}</span></td>
      <td><span class="after">{self._esc(f.after)}</span></td>
    </tr>"""

        return f"""
<div class="section">
  <div class="section-header {header_class}">{title} ({len(fixes)})</div>
  <table class="fix-table">
    <tr><th>规则</th><th>Sheet</th><th>用例</th><th>步骤</th><th>修改前</th><th>修改后</th></tr>
    {rows}
  </table>
</div>"""

    def _render_warn_section(self, title: str,
                             fixes: List[FixRecord]) -> str:
        if not fixes:
            return f"""
<div class="section">
  <div class="section-header header-warn">{title} (0)</div>
  <div class="empty-msg">无此类问题 ✓</div>
</div>"""

        rows = ""
        for f in fixes:
            rows += f"""
    <tr>
      <td><span class="rule-tag">{f.rule_id}</span></td>
      <td>{f.sheet}</td>
      <td>{f.case_name}</td>
      <td>{f.step_num}</td>
      <td>{self._esc(f.before)}</td>
      <td><span class="after">{self._esc(f.after)}</span></td>
    </tr>"""

        return f"""
<div class="section">
  <div class="section-header header-warn">{title} ({len(fixes)})</div>
  <table class="fix-table">
    <tr><th>规则</th><th>Sheet</th><th>用例</th><th>步骤</th><th>当前描述</th><th>建议</th></tr>
    {rows}
  </table>
</div>"""

    def _render_l3_error_section(self, title: str,
                                fixes: List[FixRecord]) -> str:
        """渲染 L3 解析错误区块（红色主题，显示原始步骤 + 修改建议）"""
        if not fixes:
            return f"""
<div class="section">
  <div class="section-header header-l3">{title} (0)</div>
  <div class="empty-msg">无 L3 解析错误 ✓</div>
</div>"""

        rows = ""
        for f in fixes:
            suggestion_html = self._esc(f.suggestion or '').replace('\n', '<br>')
            rows += f"""
    <tr>
      <td><span class="rule-tag">{f.rule_id}</span></td>
      <td>{f.sheet}</td>
      <td>{f.case_name}</td>
      <td>{f.step_num}</td>
      <td><span class="before">{self._esc(f.before)}</span></td>
      <td><span class="suggestion">{suggestion_html}</span></td>
    </tr>"""

        return f"""
<div class="section">
  <div class="section-header header-l3">{title} ({len(fixes)})</div>
  <table class="fix-table">
    <tr><th>规则</th><th>Sheet</th><th>用例</th><th>步骤</th><th>当前描述</th><th>修改建议</th></tr>
    {rows}
  </table>
</div>"""

    def _esc(self, text: str) -> str:
        """HTML 转义"""
        if not text:
            return ''
        return (text.replace('&', '&amp;')
                .replace('<', '&lt;')
                .replace('>', '&gt;')
                .replace('"', '&quot;'))


# ─── CLI ───

def main():
    parser = argparse.ArgumentParser(
        description='Excel 测试用例预检工具 (Phase 1)')
    parser.add_argument('input', help='输入 Excel 文件路径')
    parser.add_argument('--output', '-o', help='修正版输出路径 (默认: 输入文件名-修正版.xlsx)')
    parser.add_argument('--report', '-r', help='HTML 报告路径 (默认: excel_validation_report.html)')

    args = parser.parse_args()
    validator = ExcelValidator(args.input, args.output, args.report)
    exit_code = validator.run()
    sys.exit(exit_code)


if __name__ == '__main__':
    main()
