#!/usr/bin/env python3
"""
read_excel.py — 多 Sheet Excel 测试用例解析器

自动适配列标题变体，输出标准化 JSON。
供 generate-ui-test Phase 0 使用，替代 AI 手动读取 Excel。

用法:
    # 读取所有 Sheet
    python read_excel.py path/to/testcases.xlsx

    # 只读取指定 Sheet
    python read_excel.py path/to/testcases.xlsx --sheets "问题管理" "工单管理"

    # 输出到文件
    python read_excel.py path/to/testcases.xlsx --output parsed_cases.json

    # 显示原始步骤文本（不拆分）
    python read_excel.py path/to/testcases.xlsx --no-split-steps

支持的列标题变体:
    用例名称 / 测试用例名称* / 测试用例名称 / Case Name
    用例步骤 / 测试用例内容* / 测试用例内容 / Step Description
    模块 / Module
"""

import argparse
import json
import os
import re
import sys

# Ensure tools/ is on sys.path for core.* imports
_tools_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _tools_dir not in sys.path:
    sys.path.insert(0, _tools_dir)

try:
    import openpyxl
except ImportError:
    print("[FATAL] 需要 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

try:
    import yaml
except ImportError:
    yaml = None


# 列标题变体映射（role -> 所有可能的标题名）
COLUMN_ALIASES = {
    'case_name': ['用例名称', '测试用例名称*', '测试用例名称', 'Case Name', 'case_name'],
    'step_desc': ['用例步骤', '测试用例内容*', '测试用例内容', 'Step Description', 'step_desc', 'steps'],
    'module':    ['模块', 'Module', 'module'],
}

# 步骤分隔正则：编号列表（1. / 1、/ 步骤1：）
STEP_SPLIT_RE = re.compile(r'(?:^|\n)\s*\d+[.、．：:]\s*')
STEP_NUMBER_RE = re.compile(r'^\s*(\d+)[.、．：:]\s*')


def detect_columns(headers: list) -> dict:
    """自动检测列标题，返回 {role: column_index} 映射"""
    mapping = {}
    for role, aliases in COLUMN_ALIASES.items():
        for i, h in enumerate(headers):
            # 精确匹配（去除首尾空格和星号）
            h_clean = str(h or '').strip().rstrip('*')
            for alias in aliases:
                alias_clean = alias.strip().rstrip('*')
                if h_clean == alias_clean:
                    mapping[role] = i
                    break
            if role in mapping:
                break
    return mapping


def split_steps(text: str) -> list:
    """将编号列表文本拆分为独立步骤

    输入: "1. 访问URL\\n2. 点击按钮\\n3. 验证结果"
    输出: ["访问URL", "点击按钮", "验证结果"]
    """
    if not text or not text.strip():
        return []

    # 按编号拆分
    parts = STEP_SPLIT_RE.split(text.strip())
    steps = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 去掉开头可能的残余编号
        part = STEP_NUMBER_RE.sub('', part).strip()
        if part:
            steps.append(part)

    # 如果没有拆分成功，返回整段文本作为单步
    if not steps:
        steps = [text.strip()]

    return steps


def parse_sheet(ws, sheet_name: str, split_steps_flag: bool = True) -> dict:
    """解析单个 Sheet，返回标准化结构"""
    headers = [str(c.value or '').strip() for c in ws[1]]
    col_map = detect_columns(headers)

    if 'case_name' not in col_map:
        return {
            'sheet': sheet_name,
            'headers': headers,
            'error': '未找到用例名称列（尝试的标题: ' +
                     ', '.join(COLUMN_ALIASES['case_name']) + '）',
            'cases': [],
        }

    if 'step_desc' not in col_map:
        return {
            'sheet': sheet_name,
            'headers': headers,
            'error': '未找到用例步骤列（尝试的标题: ' +
                     ', '.join(COLUMN_ALIASES['step_desc']) + '）',
            'cases': [],
        }

    name_idx = col_map['case_name']
    steps_idx = col_map['step_desc']
    module_idx = col_map.get('module')

    cases = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过空行
        case_name = str(row[name_idx] or '').strip() if name_idx < len(row) else ''
        if not case_name:
            continue

        steps_text = str(row[steps_idx] or '').strip() if steps_idx < len(row) else ''
        module_name = str(row[module_idx] or '').strip() if module_idx is not None and module_idx < len(row) else ''

        case = {
            'row': row_num,
            'module': module_name,
            'case_name': case_name,
        }

        if split_steps_flag:
            case['steps'] = split_steps(steps_text)
            case['step_count'] = len(case['steps'])
        else:
            case['steps_raw'] = steps_text

        cases.append(case)

    return {
        'sheet': sheet_name,
        'headers': headers,
        'column_mapping': {k: headers[v] for k, v in col_map.items() if v < len(headers)},
        'cases': cases,
    }


def read_excel(filepath: str, sheets: list = None, split_steps_flag: bool = True) -> list:
    """读取 Excel 文件的所有（或指定）Sheet，返回标准化结果列表"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    target_sheets = sheets or wb.sheetnames
    results = []

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            results.append({
                'sheet': sheet_name,
                'error': f'Sheet "{sheet_name}" 不存在（可用: {", ".join(wb.sheetnames)}）',
                'cases': [],
            })
            continue

        ws = wb[sheet_name]
        result = parse_sheet(ws, sheet_name, split_steps_flag)
        results.append(result)

    wb.close()
    return results


def print_summary(results: list):
    """打印解析摘要"""
    total_cases = 0
    for r in results:
        case_count = len(r.get('cases', []))
        total_cases += case_count
        status = 'OK' if 'error' not in r else f'ERR ({r["error"]})'
        col_info = ''
        if 'column_mapping' in r:
            col_info = ' | 列: ' + ', '.join(f'{k}→{v}' for k, v in r['column_mapping'].items())
        print(f"  {r['sheet']}: {case_count} 条用例 {status}{col_info}")

    print(f"\n  总计: {total_cases} 条用例（{len(results)} 个 Sheet）")


def extract_labels_from_excel(filepath: str, output_path: str) -> int:
    """从 Excel 提取每个模块的操作对象白名单，输出 JSON

    复用 step_patterns.py 的模式匹配，提取步骤中的操作对象文本。
    用于 probe_element.py --whitelist 参数，防止 AI 编造不存在的标签。
    """
    # 导入 step_patterns（单一真相源）
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    try:
        from core.step_patterns import STEP_PATTERNS
    except ImportError:
        print("[WARN] 无法导入 step_patterns，使用内置正则", file=sys.stderr)
        STEP_PATTERNS = []

    # 非 step_patterns 的补充正则（覆盖 step_patterns 未覆盖的操作对象提取）
    _SUPPLEMENT_PATTERNS = [
        (re.compile(r'选择\s*["“](.+?)["”]\s*(?:下拉框|下拉|选项)'), 'field'),
        (re.compile(r'输入\s*["“](.+?)["”]'), 'field'),
        (re.compile(r'勾选\s*["“](.+?)["”]'), 'field'),
    ]

    # 排除的 action_type（不是操作对象）
    _EXCLUDE_TYPES = {'assert', 'assert_row', 'check_assert', 'wait',
                      'log', 'open_url', 'screenshot'}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    whitelist = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value or '').strip() for c in ws[1]]
        col_map = detect_columns(headers)

        if 'step_desc' not in col_map:
            continue

        steps_idx = col_map['step_desc']
        label_map = {}  # label -> {source_steps}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if steps_idx >= len(row) or not row[steps_idx]:
                continue
            steps_text = str(row[steps_idx]).strip()
            steps = split_steps(steps_text)

            for step in steps:
                label = None

                # 优先用 step_patterns 匹配
                for compiled_re, action_type, group_names in STEP_PATTERNS:
                    if action_type in _EXCLUDE_TYPES:
                        continue
                    m = compiled_re.search(step)
                    if m and m.groups():
                        label = m.group(1).strip()
                        break

                # step_patterns 未匹配，尝试补充正则
                if not label:
                    for pat, _ in _SUPPLEMENT_PATTERNS:
                        m = pat.search(step)
                        if m:
                            label = m.group(1).strip()
                            break

                if not label or len(label) < 2:
                    continue
                # 过滤掉纯数字或特殊字符
                if label[0].isdigit() or all(c in '.,;:!?。，；：！？' for c in label):
                    continue

                if label not in label_map:
                    label_map[label] = {'label': label, 'source_steps': []}
                if len(label_map[label]['source_steps']) < 3:
                    label_map[label]['source_steps'].append(step[:80])

        labels_list = list(label_map.values())
        whitelist[sheet_name] = {
            'sheet_name': sheet_name,
            'labels': labels_list,
            'label_set': sorted(label_map.keys()),
            'total': len(labels_list),
        }

    wb.close()

    # 输出
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(whitelist, f, ensure_ascii=False, indent=2)

    print(f"[OK] 白名单已生成: {output_path}", file=sys.stderr)
    for name, data in whitelist.items():
        print(f"  {name}: {data['total']} 个标签", file=sys.stderr)
    return 0


# ============================================================================
# --extract-urls: 从 Excel 提取模块 URL 映射
# ============================================================================

_STANDARD_RE = re.compile(r'^#\s*模块[:：]\s*(.+)$')
_TRAILING_RE = re.compile(r'^#\s*(.+?)模块\s*$')


def extract_urls_from_excel(filepath, output_path, pages_dir=None, config_path=None):
    """从 Excel 按第一列模块值分组收集 URL，不翻译模块名。

    修复: Issue 1 — 直接用中文名作 key，永不因翻译失败而丢失 URL。
    翻译逻辑由独立工具 build_module_map.py 完成，输出 _probe/module_map.json 供下游使用。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    url_pattern = re.compile(r'访问\s*(https?://\S+)')

    # {cn_module: set(urls)}
    module_urls = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value or '').strip() for c in ws[1]]
        col_map = detect_columns(headers)

        if 'step_desc' not in col_map:
            continue
        steps_idx = col_map['step_desc']
        module_idx = col_map.get('module')

        # 无"模块"列时回退到 sheet 名称（而非跳过整个 sheet）
        fallback_module = sheet_name if module_idx is None else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if module_idx is not None:
                if module_idx >= len(row):
                    continue
                cn_module = str(row[module_idx] or '').strip()
            else:
                cn_module = fallback_module
            if not cn_module:
                continue

            steps_text = str(row[steps_idx]).strip() if steps_idx < len(row) and row[steps_idx] else ''
            for step in split_steps(steps_text):
                m = url_pattern.search(step)
                if m:
                    url = m.group(1)
                    module_urls.setdefault(cn_module, set()).add(url)
                    break  # 每个 case 只取第一个 URL

    wb.close()

    # 输出：直接用中文名作 key
    result = {cn: {"urls": sorted(urls)} for cn, urls in sorted(module_urls.items())}

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    total_urls = sum(len(v['urls']) for v in result.values())
    print(f"[OK] module_urls.json: {len(result)} 个模块, {total_urls} 个唯一 URL",
          file=sys.stderr)
    for cn, data in result.items():
        print(f"  {cn}: {len(data['urls'])} URLs",
              file=sys.stderr)
    return 0


def main():
    parser = argparse.ArgumentParser(
        description='多 Sheet Excel 测试用例解析器（自动适配列标题变体）'
    )
    parser.add_argument('filepath', help='Excel 文件路径（.xlsx）')
    parser.add_argument('--sheets', nargs='*', help='只读取指定 Sheet（默认全部）')
    parser.add_argument('--output', help='输出 JSON 文件路径（默认输出到 stdout）')
    parser.add_argument('--no-split-steps', action='store_true',
                        help='不拆分步骤，保留原始步骤文本')
    parser.add_argument('--summary', action='store_true',
                        help='仅显示摘要，不输出完整 JSON')
    parser.add_argument('--extract-labels', action='store_true',
                        help='提取操作对象白名单（输出 JSON 供 probe --whitelist 使用）')
    parser.add_argument('--extract-urls', action='store_true',
                        help='提取模块 URL 映射（输出 JSON）')
    parser.add_argument('--pages-dir', default=None,
                        help='pages/ 目录路径（用于 YAML 注释扫描）')
    parser.add_argument('--config', default=None,
                        help='config.yaml 路径（用于 module_aliases 读取）')

    args = parser.parse_args()

    # Windows 终端编码兼容
    if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')
    if sys.stderr.encoding and sys.stderr.encoding.lower() != 'utf-8':
        sys.stderr.reconfigure(encoding='utf-8')

    # --extract-labels 和 --extract-urls 互斥检查
    if args.extract_labels and args.extract_urls:
        print("[ERROR] --extract-labels 和 --extract-urls 不能同时使用", file=sys.stderr)
        sys.exit(1)

    # --extract-urls 模式：提取 URL 映射后直接退出
    if args.extract_urls:
        if not args.output:
            print("[ERROR] --extract-urls 需要 --output 参数", file=sys.stderr)
            sys.exit(1)
        sys.exit(extract_urls_from_excel(
            args.filepath, args.output,
            pages_dir=args.pages_dir, config_path=args.config))

    # --extract-labels 模式：提取白名单后直接退出
    if args.extract_labels:
        if not args.output:
            print("[ERROR] --extract-labels 需要 --output 参数", file=sys.stderr)
            sys.exit(1)
        sys.exit(extract_labels_from_excel(args.filepath, args.output))

    results = read_excel(args.filepath, args.sheets, not args.no_split_steps)

    if args.summary:
        print_summary(results)
        return

    # 输出 JSON
    output = json.dumps(results, ensure_ascii=False, indent=2)

    if args.output:
        os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(output)
        print(f"[OK] 已输出到: {args.output}", file=sys.stderr)
        print_summary(results)
    else:
        print(output)


if __name__ == '__main__':
    main()
