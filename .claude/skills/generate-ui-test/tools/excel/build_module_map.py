#!/usr/bin/env python3
"""build_module_map.py — 构建中文模块名→英文 slug 映射（Phase 2.5）

从 Excel 模块列 + pages/ 目录结构 + YAML 注释 + discovery JSON 自动构建映射，
输出 _probe/module_map.json 供 Phase 4、Phase 5 使用。

用法:
    python build_module_map.py "{excel_file}" \
      --pages {project}/pages \
      --discovery-dir {project}/_probe \
      --output {project}/_probe/module_map.json \
      --module-map "总览查看=overview-mail,站内信查看=overview-mail"  # 可选覆盖
"""
import argparse
import glob
import json
import os
import re
import sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("[ERROR] 缺少 openpyxl，请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 复用 read_excel.py 的列检测逻辑
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
try:
    from read_excel import detect_columns, COLUMN_ALIASES
except ImportError as e:
    print(f"[ERROR] 无法导入 read_excel.py: {e}", file=sys.stderr)
    print("[INFO] 请确保 read_excel.py 存在于 tools/ 目录", file=sys.stderr)
    sys.exit(1)


def _extract_excel_modules(excel_path):
    """从 Excel 提取所有唯一的中文模块名（模块列值）。

    遍历所有 sheet，检测"模块"列（自动适配列标题变体），收集非空值。
    如果 sheet 没有"模块"列，回退到 sheet 名称（与 read_excel.py extract_urls 修改7b 一致）。
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    modules = set()
    for ws in wb.worksheets:
        headers = [str(c.value or '').strip() for c in ws[1]]
        col_map = detect_columns(headers)
        module_idx = col_map.get('module')
        if module_idx is None:
            # 回退到 sheet 名称（与 read_excel.py extract_urls 修改7b 一致）
            sheet_name = ws.title
            if sheet_name and sheet_name != 'Sheet':
                modules.add(sheet_name)
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if module_idx >= len(row):
                continue
            val = row[module_idx]
            if val and isinstance(val, str):
                name = val.strip()
                if name and name != '模块':
                    modules.add(name)
    wb.close()
    return sorted(modules)


def _scan_pages_dirs(pages_dir):
    """扫描 pages/ 目录，返回英文 slug 集合。

    过滤规则：
    - 排除 '.' 和 '_' 开头的目录
    - 排除包含中文字符的目录（历史污染数据）
    - 排除 'common' 目录（通用元素）
    """
    slugs = set()
    if not os.path.isdir(pages_dir):
        return slugs
    for entry in os.listdir(pages_dir):
        full = os.path.join(pages_dir, entry)
        # 基本过滤
        if not os.path.isdir(full):
            continue
        if entry.startswith(('.', '_')):
            continue
        if entry == 'common':
            continue
        # 过滤中文目录名（包含非ASCII字符）
        if any(ord(c) > 127 for c in entry):
            continue
        slugs.add(entry)
    return slugs


def _scan_yaml_comments(pages_dir):
    """扫描 pages/*/elements.yaml 注释，提取 `# 模块: XX` 映射。

    Returns: {中文名: slug}
    """
    mapping = {}
    if not os.path.isdir(pages_dir):
        return mapping
    for slug in os.listdir(pages_dir):
        yaml_path = os.path.join(pages_dir, slug, 'elements.yaml')
        if not os.path.isfile(yaml_path):
            continue
        try:
            with open(yaml_path, encoding='utf-8') as f:
                for line in f:
                    m = re.match(r'^#\s*模块:\s*(.+)', line)
                    if m:
                        cn_name = m.group(1).strip()
                        mapping[cn_name] = slug
                        break
        except Exception:
            continue
    return mapping


def _scan_discovery_json(discovery_dir):
    """扫描 _probe/discovery_*.json，提取 cn_name 字段映射。

    Returns: {cn_name: slug}
    """
    mapping = {}
    if not os.path.isdir(discovery_dir):
        return mapping
    for disc_file in sorted(glob.glob(os.path.join(discovery_dir, 'discovery_*.json'))):
        try:
            with open(disc_file, encoding='utf-8') as f:
                disc = json.load(f)
        except Exception:
            continue
        slug = disc.get('module', '')
        cn = disc.get('cn_name', '')
        if cn and slug and cn != slug:
            mapping[cn] = slug
    return mapping


def _auto_generate_slug(cn_name):
    """从中文模块名自动生成英文 slug（首次运行兜底）。

    策略链（与 generate_from_excel.py _auto_generate_slug_inline 完全一致）:
      1. 提取 ASCII 部分（如 "PMO管理" → "pmo"），长度 >= 3 才采用
      2. MD5 hash 兜底（mod_ 前缀 + 8 位 hex）

    Returns: 合法 slug 字符串（永不为 None/空，确保首次运行不阻断）。
    """
    # 策略 1: ASCII 部分
    ascii_part = re.sub(r'[^a-zA-Z0-9]', '', cn_name).lower()
    if len(ascii_part) >= 3:
        return ascii_part

    # 策略 2: MD5 hash 兜底
    import hashlib
    return 'mod_' + hashlib.md5(cn_name.encode('utf-8')).hexdigest()[:8]


def _extract_slug_from_url(cn_name, module_urls):
    """从 module_urls.json 的 URL 路径提取 slug。

    提取策略（按优先级）:
    1. Hash 路由 #/path 的第一个段
    2. 普通路径的第三个段（如果前两个段是通用名称如 /estack/web/）
    3. 普通路径的第二个段（如果第一个段是通用名称）
    4. 普通路径的第一个段

    Args:
        cn_name: 中文模块名
        module_urls: module_urls.json 内容 {中文名: {urls: [...]}}

    Returns: slug 字符串或 None
    """
    if not module_urls or cn_name not in module_urls:
        return None
    urls = module_urls[cn_name].get('urls', [])
    if not urls:
        return None
    url = urls[0]

    def _validate_segment(seg):
        """校验 segment 是否为合法 slug"""
        return seg and re.match(r'^[a-z][a-z0-9_-]{1,29}$', seg)

    def _is_generic_segment(seg):
        """判断 segment 是否为通用名称（不适合作为模块名）"""
        generic_names = {'estack', 'web', 'api', 'static', 'app', 'console'}
        return seg.lower() in generic_names

    # 策略 1: 优先匹配 #/path 格式（hash 路由）
    match = re.search(r'#/([^/]+)', url)
    if match:
        segment = match.group(1).replace('-', '_')  # 连字符转下划线
        if _validate_segment(segment):
            return segment

    # 提取普通路径的所有段
    match = re.search(r'https?://[^/]+(/[^?#]*)', url)
    if match:
        path = match.group(1)
        segments = [s for s in path.split('/') if s]

        # 策略 2: 尝试第三个段（如果前两个是通用名称）
        if len(segments) >= 3:
            if _is_generic_segment(segments[0]) and _is_generic_segment(segments[1]):
                segment = segments[2]
                if _validate_segment(segment) and not _is_generic_segment(segment):
                    return segment

        # 策略 3: 尝试第二个段（如果第一个是通用名称）
        if len(segments) >= 2:
            if _is_generic_segment(segments[0]):
                segment = segments[1]
                if _validate_segment(segment) and not _is_generic_segment(segment):
                    return segment

        # 策略 4: 第一个段（兜底）
        if segments:
            segment = segments[0]
            if _validate_segment(segment):
                return segment

    return None


def _build_mapping(cn_modules, en_slugs, yaml_comments, discovery_map, cli_overrides, module_urls=None):
    """构建中文模块名→英文 slug 映射。

    匹配优先级:
    1. CLI --module-map 显式覆盖
    2. discovery JSON 的 cn_name 字段（历史兼容）
    3. YAML 注释精确匹配 (cn_name == comment)
    4. YAML 注释子串匹配 (cn_name ⊂ comment 或 comment ⊂ cn_name)
    5. 中文名本身是英文 slug (cn_name in en_slugs)
    6. URL 路径第一段提取（如 #/instation-mail/... → instation-mail）
    7. 自动 slug 生成（ASCII提取/MD5兜底）
    """
    result = OrderedDict()
    unmatched = []

    for cn in cn_modules:
        # Priority 1: CLI override
        if cn in cli_overrides:
            result[cn] = cli_overrides[cn]
            continue

        # Priority 2: discovery JSON cn_name
        if cn in discovery_map:
            result[cn] = discovery_map[cn]
            continue

        # Priority 3: YAML comment exact match
        if cn in yaml_comments:
            result[cn] = yaml_comments[cn]
            continue

        # Priority 4: YAML comment substring match
        matched = False
        sorted_comments = sorted(yaml_comments.items(),
                                 key=lambda x: len(x[0]), reverse=True)
        for comment, slug in sorted_comments:
            if cn in comment or comment in cn:
                result[cn] = slug
                matched = True
                break
        if matched:
            continue

        # Priority 5: cn_name is already an en_slug
        if cn in en_slugs:
            result[cn] = cn
            continue

        # Priority 6: URL path segment extraction
        if module_urls:
            url_slug = _extract_slug_from_url(cn, module_urls)
            if url_slug:
                # 碰撞检测
                if url_slug in result.values():
                    original = url_slug
                    for suffix in range(2, 100):
                        candidate = f'{original}-{suffix}'
                        if candidate not in result.values():
                            url_slug = candidate
                            break
                    print(f"[URL-SLUG] 碰撞检测: {cn} → {url_slug} (原值 {original} 与已有模块冲突)",
                          file=sys.stderr)
                result[cn] = url_slug
                print(f"[URL-SLUG] {cn} → {url_slug}（从 URL 路径提取）",
                      file=sys.stderr)
                continue

        # Priority 7: 自动 slug 生成（首次运行兜底，永不阻断）
        auto_slug = _auto_generate_slug(cn)
        if auto_slug:
            # 碰撞检测：如果 auto_slug 已存在（不同 cn 生成相同 hash），追加后缀
            if auto_slug in result.values():
                original = auto_slug
                for suffix in range(2, 100):
                    candidate = f'{original}_{suffix}'
                    if candidate not in result.values():
                        auto_slug = candidate
                        break
                print(f"[AUTO-SLUG] 碰撞检测: {cn} → {auto_slug} (原值 {original} 与已有模块冲突)",
                      file=sys.stderr)
            result[cn] = auto_slug
            print(f"[AUTO-SLUG] {cn} → {auto_slug}（自动生成，建议用 --module-map 确认）",
                  file=sys.stderr)
            continue

        # Priority 7: unmatched
        unmatched.append(cn)

    if unmatched:
        print(f"[ERROR] 以下中文模块名无法匹配到英文 slug:", file=sys.stderr)
        for name in unmatched:
            print(f"  - {name}", file=sys.stderr)
        print(f"\n已匹配的模块:", file=sys.stderr)
        for cn, slug in result.items():
            print(f"  {cn} → {slug}", file=sys.stderr)
        print(f"\n可用的英文 slug: {sorted(en_slugs)}", file=sys.stderr)
        print(f"可用的 YAML 注释: {list(yaml_comments.keys())}", file=sys.stderr)
        print(f"可用的 discovery cn_name: {list(discovery_map.keys())}", file=sys.stderr)
        print(f"\n请使用 --module-map 参数手动指定未匹配的模块映射:", file=sys.stderr)
        mapping_str = ','.join(f'{name}=slug' for name in unmatched)
        print(f'  --module-map "{mapping_str}"', file=sys.stderr)
        sys.exit(1)

    return result


def main():
    parser = argparse.ArgumentParser(
        description='构建中文模块名→英文 slug 映射文件')
    parser.add_argument('excel', help='Excel 用例文件路径')
    parser.add_argument('--pages', required=True,
                        help='pages/ 目录路径')
    parser.add_argument('--discovery-dir', default=None,
                        help='_probe/ 目录路径（可选，兼容历史 discovery JSON）')
    parser.add_argument('--output', required=True,
                        help='输出 JSON 文件路径')
    parser.add_argument('--module-map', default='',
                        help='手动覆盖映射，格式: 中文名1=slug1,中文名2=slug2')
    parser.add_argument('--module-urls', default=None,
                        help='module_urls.json 路径（可选，用于从 URL 提取 slug）')

    args = parser.parse_args()

    # 验证输入
    if not os.path.isfile(args.excel):
        print(f"[ERROR] Excel 文件不存在: {args.excel}", file=sys.stderr)
        sys.exit(1)

    # 解析 CLI 覆盖
    cli_overrides = {}
    if args.module_map:
        for pair in args.module_map.split(','):
            pair = pair.strip()
            if '=' in pair:
                k, v = pair.split('=', 1)
                cli_overrides[k.strip()] = v.strip()

    # Step 1: 提取 Excel 中文模块名
    cn_modules = _extract_excel_modules(args.excel)
    print(f"[INFO] Excel 中发现 {len(cn_modules)} 个中文模块名: {cn_modules}")

    # Step 2: 扫描 pages/ 英文 slug
    en_slugs = _scan_pages_dirs(args.pages)
    print(f"[INFO] pages/ 中发现 {len(en_slugs)} 个英文 slug: {sorted(en_slugs)}")

    # Step 3: 扫描 YAML 注释
    yaml_comments = _scan_yaml_comments(args.pages)
    print(f"[INFO] YAML 注释映射: {yaml_comments}")

    # Step 4: 扫描 discovery JSON cn_name（可选）
    discovery_map = {}
    if args.discovery_dir:
        discovery_map = _scan_discovery_json(args.discovery_dir)
        print(f"[INFO] discovery JSON cn_name 映射: {discovery_map}")

    # Step 4.5: 加载 module_urls.json（可选，用于 URL 路径提取）
    module_urls = None
    if args.module_urls and os.path.isfile(args.module_urls):
        try:
            with open(args.module_urls, encoding='utf-8') as f:
                module_urls = json.load(f)
            print(f"[INFO] 已加载 module_urls.json: {len(module_urls)} 个模块")
        except Exception as e:
            print(f"[WARN] 无法加载 module_urls.json: {e}", file=sys.stderr)

    # Step 5: 构建映射
    mapping = _build_mapping(cn_modules, en_slugs, yaml_comments, discovery_map, cli_overrides, module_urls)

    # Step 6: 输出
    os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)

    print(f"\n[OK] 模块映射已写入: {args.output}")
    for cn, slug in mapping.items():
        print(f"  {cn} → {slug}")


if __name__ == '__main__':
    main()
