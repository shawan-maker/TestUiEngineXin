#!/usr/bin/env python3
"""build_module_map.py — 构建中文模块名→英文 slug 映射（Phase 2.5）

从 Excel 模块列 + pages/ 目录结构 + YAML 注释 + discovery JSON 自动构建映射，
输出 _probe/module_map.json 供 Phase 4、Phase 5 使用。

用法:
    python build_module_map.py "{excel_file}" \
      --pages {project}/pages \
      --discovery-dir {project}/_probe \
      --output {project}/_probe/module_map.json \
      --module-map "总览查看=overview-mail,站内信查看=overview-mail"  # 可选覆盖
"""
import argparse
import glob
import json
import os
import re
import sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("[ERROR] 缺少 openpyxl，请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 复用 read_excel.py 的列检测逻辑
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
try:
    from read_excel import detect_columns, COLUMN_ALIASES
except ImportError as e:
    print(f"[ERROR] 无法导入 read_excel.py: {e}", file=sys.stderr)
    print("[INFO] 请确保 read_excel.py 存在于 tools/ 目录", file=sys.stderr)
    sys.exit(1)


def _extract_excel_modules(excel_path):
    """从 Excel 提取所有唯一的中文模块名（模块列值）。

    遍历所有 sheet，检测"模块"列（自动适配列标题变体），收集非空值。
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    modules = set()
    for ws in wb.worksheets:
        headers = [str(c.value or '').strip() for c in ws[1]]
        col_map = detect_columns(headers)
        module_idx = col_map.get('module')
        if module_idx is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if module_idx >= len(row):
                continue
            val = row[module_idx]
            if val and isinstance(val, str):
                name = val.strip()
                if name and name != '模块':
                    modules.add(name)
    wb.close()
    return sorted(modules)


def _scan_pages_dirs(pages_dir):
    """扫描 pages/ 目录，返回英文 slug 集合。

    过滤规则：
    - 排除 '.' 和 '_' 开头的目录
    - 排除包含中文字符的目录（历史污染数据）
    - 排除 'common' 目录（通用元素）
    """
    slugs = set()
    if not os.path.isdir(pages_dir):
        return slugs
    for entry in os.listdir(pages_dir):
        full = os.path.join(pages_dir, entry)
        # 基本过滤
        if not os.path.isdir(full):
            continue
        if entry.startswith(('.', '_')):
            continue
        if entry == 'common':
            continue
        # 过滤中文目录名（包含非ASCII字符）
        if any(ord(c) > 127 for c in entry):
            continue
        slugs.add(entry)
    return slugs


def _scan_yaml_comments(pages_dir):
    """扫描 pages/*/elements.yaml 注释，提取 `# 模块: XX` 映射。

    Returns: {中文名: slug}
    """
    mapping = {}
    if not os.path.isdir(pages_dir):
        return mapping
    for slug in os.listdir(pages_dir):
        yaml_path = os.path.join(pages_dir, slug, 'elements.yaml')
        if not os.path.isfile(yaml_path):
            continue
        try:
            with open(yaml_path, encoding='utf-8') as f:
                for line in f:
                    m = re.match(r'^#\s*模块:\s*(.+)', line)
                    if m:
                        cn_name = m.group(1).strip()
                        mapping[cn_name] = slug
                        break
        except Exception:
            continue
    return mapping


def _scan_discovery_json(discovery_dir):
    """扫描 _probe/discovery_*.json，提取 cn_name 字段映射。

    Returns: {cn_name: slug}
    """
    mapping = {}
    if not os.path.isdir(discovery_dir):
        return mapping
    for disc_file in sorted(glob.glob(os.path.join(discovery_dir, 'discovery_*.json'))):
        try:
            with open(disc_file, encoding='utf-8') as f:
                disc = json.load(f)
        except Exception:
            continue
        slug = disc.get('module', '')
        cn = disc.get('cn_name', '')
        if cn and slug and cn != slug:
            mapping[cn] = slug
    return mapping


def _build_mapping(cn_modules, en_slugs, yaml_comments, discovery_map, cli_overrides):
    """构建中文模块名→英文 slug 映射。

    匹配优先级:
    1. CLI --module-map 显式覆盖
    2. discovery JSON 的 cn_name 字段（历史兼容）
    3. YAML 注释精确匹配 (cn_name == comment)
    4. YAML 注释子串匹配 (cn_name ⊂ comment 或 comment ⊂ cn_name)
    5. 中文名本身是英文 slug (cn_name in en_slugs)
    6. 未匹配 → 报错退出
    """
    result = OrderedDict()
    unmatched = []

    for cn in cn_modules:
        # Priority 1: CLI override
        if cn in cli_overrides:
            result[cn] = cli_overrides[cn]
            continue

        # Priority 2: discovery JSON cn_name
        if cn in discovery_map:
            result[cn] = discovery_map[cn]
            continue

        # Priority 3: YAML comment exact match
        if cn in yaml_comments:
            result[cn] = yaml_comments[cn]
            continue

        # Priority 4: YAML comment substring match
        matched = False
        sorted_comments = sorted(yaml_comments.items(),
                                 key=lambda x: len(x[0]), reverse=True)
        for comment, slug in sorted_comments:
            if cn in comment or comment in cn:
                result[cn] = slug
                matched = True
                break
        if matched:
            continue

        # Priority 5: cn_name is already an en_slug
        if cn in en_slugs:
            result[cn] = cn
            continue

        # Priority 6: unmatched
        unmatched.append(cn)

    if unmatched:
        print(f"[ERROR] 以下中文模块名无法匹配到英文 slug:", file=sys.stderr)
        for name in unmatched:
            print(f"  - {name}", file=sys.stderr)
        print(f"\n已匹配的模块:", file=sys.stderr)
        for cn, slug in result.items():
            print(f"  {cn} → {slug}", file=sys.stderr)
        print(f"\n可用的英文 slug: {sorted(en_slugs)}", file=sys.stderr)
        print(f"可用的 YAML 注释: {list(yaml_comments.keys())}", file=sys.stderr)
        print(f"可用的 discovery cn_name: {list(discovery_map.keys())}", file=sys.stderr)
        print(f"\n请使用 --module-map 参数手动指定未匹配的模块映射:", file=sys.stderr)
        mapping_str = ','.join(f'{name}=slug' for name in unmatched)
        print(f'  --module-map "{mapping_str}"', file=sys.stderr)
        sys.exit(1)

    return result


def main():
    parser = argparse.ArgumentParser(
        description='构建中文模块名→英文 slug 映射文件')
    parser.add_argument('excel', help='Excel 用例文件路径')
    parser.add_argument('--pages', required=True,
                        help='pages/ 目录路径')
    parser.add_argument('--discovery-dir', default=None,
                        help='_probe/ 目录路径（可选，兼容历史 discovery JSON）')
    parser.add_argument('--output', required=True,
                        help='输出 JSON 文件路径')
    parser.add_argument('--module-map', default='',
                        help='手动覆盖映射，格式: 中文名1=slug1,中文名2=slug2')

    args = parser.parse_args()

    # 验证输入
    if not os.path.isfile(args.excel):
        print(f"[ERROR] Excel 文件不存在: {args.excel}", file=sys.stderr)
        sys.exit(1)

    # 解析 CLI 覆盖
    cli_overrides = {}
    if args.module_map:
        for pair in args.module_map.split(','):
            pair = pair.strip()
            if '=' in pair:
                k, v = pair.split('=', 1)
                cli_overrides[k.strip()] = v.strip()

    # Step 1: 提取 Excel 中文模块名
    cn_modules = _extract_excel_modules(args.excel)
    print(f"[INFO] Excel 中发现 {len(cn_modules)} 个中文模块名: {cn_modules}")

    # Step 2: 扫描 pages/ 英文 slug
    en_slugs = _scan_pages_dirs(args.pages)
    print(f"[INFO] pages/ 中发现 {len(en_slugs)} 个英文 slug: {sorted(en_slugs)}")

    # Step 3: 扫描 YAML 注释
    yaml_comments = _scan_yaml_comments(args.pages)
    print(f"[INFO] YAML 注释映射: {yaml_comments}")

    # Step 4: 扫描 discovery JSON cn_name（可选）
    discovery_map = {}
    if args.discovery_dir:
        discovery_map = _scan_discovery_json(args.discovery_dir)
        print(f"[INFO] discovery JSON cn_name 映射: {discovery_map}")

    # Step 5: 构建映射
    mapping = _build_mapping(cn_modules, en_slugs, yaml_comments, discovery_map, cli_overrides)

    # Step 6: 输出
    os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)

    print(f"\n[OK] 模块映射已写入: {args.output}")
    for cn, slug in mapping.items():
        print(f"  {cn} → {slug}")


if __name__ == '__main__':
    main()
